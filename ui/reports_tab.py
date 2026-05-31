"""
Reports Tab - Phase 3

Responsibilities:
- Run FIFO engine on tab open (recalculate every time)
- Calculate realized P/L
- Display summary cards (Total Profit, Total Loss, Net P/L)
- Show emotion indicator based on net P/L (loss/profit)
- Display daily/monthly aggregations
- Show detailed P/L breakdown

Does NOT:
- Cache results (always fresh calculation)
- Edit trades
- Validate data integrity
"""

import csv
import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
from typing import Callable, Sequence
from datetime import datetime
from pathlib import Path

import config
from core.logger import get_logger
from core.fifo_matcher import fetch_active_trades, match_fifo
from core.pnl_calculator import calculate_match_pnl
from core.mtf_interest import apply_mtf_interest
from core.pnl_aggregator import (
    aggregate_pnl_by_date,
    aggregate_pnl_by_week,
    aggregate_pnl_by_month,
    aggregate_pnl_by_year,
    aggregate_pnl_by_sell,
    aggregate_pnl_by_equity,
    aggregate_trade_value_by_date,
    filter_matches_by_date_range
)
from core.open_positions import calculate_open_positions, get_unique_equities, OpenPosition
from core.run_ledger import build_trades_by_id
from core.utils import format_money, format_money_abs, format_period_label
from core.allocations import round_divide

logger = get_logger('ui.reports_tab')


class Tooltip:
    """Lightweight tooltip for Tk widgets."""

    def __init__(self, widget: tk.Widget, text: str) -> None:
        self.widget = widget
        self.text = text
        self.tip_window: tk.Toplevel | None = None
        self.widget.bind("<Enter>", self._show)
        self.widget.bind("<Leave>", self._hide)

    def _show(self, _event: tk.Event) -> None:
        if self.tip_window or not self.text:
            return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + 20
        self.tip_window = tk.Toplevel(self.widget)
        self.tip_window.wm_overrideredirect(True)
        self.tip_window.wm_geometry(f"+{x}+{y}")
        label = ttk.Label(
            self.tip_window,
            text=self.text,
            background="#f4f6f8",
            foreground="#2c3e50",
            relief="solid",
            borderwidth=1,
            padding=(6, 3)
        )
        label.pack()

    def _hide(self, _event: tk.Event) -> None:
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None


class ReportsTab:
    """Reports tab - display P/L analysis with FIFO-based calculations."""

    @staticmethod
    def build_report_rows(period_totals: dict, mode: str) -> list[dict]:
        """
        Normalize period data into display-ready rows.

        Args:
            period_totals: Dict of period_key -> {profit, loss, net}
            mode: Period type for label formatting

        Returns:
            List of row dicts with all display values pre-calculated
        """
        rows = []
        accumulated = 0

        for period_key in sorted(period_totals.keys()):
            profit = period_totals[period_key]['profit']
            loss = period_totals[period_key]['loss']
            net = period_totals[period_key]['net']

            accumulated += net

            rows.append({
                'period_key': period_key,
                'label': format_period_label(period_key, mode),
                'profit': profit,
                'loss': loss,
                'net': net,
                'accumulated': accumulated
            })

        return rows

    def _convert_to_pnl_breakdown(self, pnl_totals: dict[str, int]) -> dict[str, dict[str, int]]:
        """
        Convert engine's simple P/L format to UI's profit/loss/net breakdown.
        Engine returns: {'2026-01-20': 1000} (net P/L in paise)
        UI needs: {'2026-01-20': {'profit': 1500, 'loss': -500, 'net': 1000}}
        """
        breakdown: dict[str, dict[str, int]] = {}
        for key, net_pnl in pnl_totals.items():
            if net_pnl > 0:
                breakdown[key] = {'profit': net_pnl, 'loss': 0, 'net': net_pnl}
            elif net_pnl < 0:
                breakdown[key] = {'profit': 0, 'loss': net_pnl, 'net': net_pnl}
            else:
                breakdown[key] = {'profit': 0, 'loss': 0, 'net': 0}
        return breakdown

    def __init__(self, parent: ttk.Frame, status_callback: Callable[[str], None]) -> None:
        logger.info("Initializing Reports tab")
        self.parent = parent
        self.update_status = status_callback

        # Data containers
        self.total_profit = 0
        self.total_loss = 0
        self.net_pnl = 0
        self.daily_pnl = {}
        self.weekly_pnl = {}
        self.monthly_pnl = {}
        self.yearly_pnl = {}

        # Filters
        self.from_date_var = tk.StringVar(value="")
        self.to_date_var = tk.StringVar(value="")
        self.show_open_positions_var = tk.BooleanVar(value=False)
        self.type1_filter_var = tk.StringVar(value="All")
        self.expiry_month_var = tk.StringVar(value="")
        self.expiry_preset_var = tk.StringVar(value="Custom")

        # Open positions / equity
        self.open_positions = []
        self.filtered_open_positions = []
        self.equity_pnl = {}
        self.analytics = {}
        self.profile_breakdown_data: dict[str, dict[str, int]] = {}
        self.audit_matches = []
        self.audit_trades_by_id = {}

        # Validation state
        self.has_validation_errors = False
        self.validation_message = ""

        # Period selection
        self.period_var = tk.StringVar(value="Daily")
        self.profile_breakdown_mode_var = tk.StringVar(value="Net P/L")

        # Create UI
        self.create_widgets()

        logger.debug("Reports tab initialized (calculations deferred)")

    def create_widgets(self) -> None:
        """Create all UI widgets for Reports tab."""
        canvas = tk.Canvas(self.parent)
        scrollbar = ttk.Scrollbar(self.parent, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        self.canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(self.canvas_window, width=e.width))
        canvas.bind_all('<MouseWheel>', lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.canvas = canvas

        main_frame = ttk.Frame(scrollable_frame, padding="20")
        main_frame.pack(fill='both', expand=True)
        self.main_frame = main_frame

        style = ttk.Style()
        style.configure("Report.TLabelframe", padding=15)
        style.configure("Report.TLabelframe.Label", font=('Consolas', 10, 'bold'))

        # Warning banner (initially hidden)
        self.warning_frame = ttk.Frame(main_frame)
        self.warning_label = ttk.Label(
            self.warning_frame,
            text="",
            font=('Consolas', 10, 'bold'),
            foreground='white',
            background='#e74c3c',
            padding=10,
            wraplength=800
        )
        self.warning_label.pack(fill='x')

        # Header
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill='x', pady=(0, 20))

        ttk.Label(
            header_frame,
            text="📊 PROFIT/LOSS REPORTS",
            font=('Consolas', 16, 'bold')
        ).pack(side='left')

        period_frame = ttk.Frame(header_frame)
        period_frame.pack(side='right', padx=20)

        ttk.Label(period_frame, text="View:", font=('Arial', 10)).pack(side='left', padx=(0, 5))
        period_dropdown = ttk.Combobox(
            period_frame,
            textvariable=self.period_var,
            values=["Daily", "Weekly", "Monthly", "Yearly"],
            state='readonly',
            width=12
        )
        period_dropdown.pack(side='left', padx=5)
        period_dropdown.bind('<<ComboboxSelected>>', lambda e: self.update_period_display())

        actions_frame = ttk.Frame(header_frame)
        actions_frame.pack(side='right')

        recalc_btn = ttk.Button(
            actions_frame,
            text="🔄 Recalculate",
            command=self.calculate_reports,
            width=15
        )
        recalc_btn.pack(side='right', padx=5)
        Tooltip(recalc_btn, "Recompute all metrics with current filters")

        self.print_btn = ttk.Button(
            actions_frame,
            text="🖨️ Print",
            command=self.print_report,
            width=12
        )
        self.print_btn.pack(side='right', padx=5)
        Tooltip(self.print_btn, "Generate a print-friendly HTML report")

        export_csv_btn = ttk.Button(
            actions_frame,
            text="⬇️ Export CSV",
            command=self.export_report_csv,
            width=14
        )
        export_csv_btn.pack(side='right', padx=5)
        Tooltip(export_csv_btn, "Export the current report to CSV")

        export_audit_btn = ttk.Button(
            actions_frame,
            text="⬇️ Audit CSV",
            command=self.export_audit_csv,
            width=14
        )
        export_audit_btn.pack(side='right', padx=5)
        Tooltip(export_audit_btn, "Export match-level audit details to CSV, including allocation remainder flags")

        export_excel_btn = ttk.Button(
            actions_frame,
            text="⬇️ Export Excel",
            command=self.export_report_excel,
            width=15
        )
        export_excel_btn.pack(side='right', padx=5)
        Tooltip(export_excel_btn, "Export the current report to Excel (xlsx)")

        # Filter Controls Section
        self.create_filter_controls(main_frame)

        # Summary Cards
        self.create_summary_cards(main_frame)

        # Analytics Section
        self.create_analytics_section(main_frame)

        ttk.Separator(main_frame, orient='horizontal').pack(fill='x', pady=20)

        # Dynamic P/L Section
        self.create_dynamic_pnl_section(main_frame)

        ttk.Separator(main_frame, orient='horizontal').pack(fill='x', pady=20)

        # Equity-wise Summary Section
        self.create_equity_summary_section(main_frame)

    def create_dynamic_pnl_section(self, parent: ttk.Frame) -> None:
        """Create dynamic P/L breakdown table that changes based on period selection."""
        context_label = ttk.Label(
            parent,
            text="Showing SELL-based Net P/L (FIFO + MTF interest)",
            font=('Consolas', 9),
            foreground='gray'
        )
        context_label.pack(anchor='w', pady=(0, 5))

        self.pnl_title_label = ttk.Label(
            parent,
            text="📅 DAILY P/L",
            font=('Consolas', 13, 'bold')
        )
        self.pnl_title_label.pack(anchor='w', pady=(0, 10))

        table_frame = ttk.Frame(parent)
        table_frame.pack(fill='both', expand=True, pady=(0, 10))

        scrollbar = ttk.Scrollbar(table_frame)
        scrollbar.pack(side='right', fill='y')

        columns = ('Period', 'Profit', 'Loss', 'Net P/L', 'Running Total')
        self.pnl_tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show='headings',
            height=12,
            yscrollcommand=scrollbar.set
        )
        scrollbar.config(command=self.pnl_tree.yview)

        self.pnl_tree.heading('Period', text='Period', anchor='w')
        self.pnl_tree.heading('Profit', text='Profit (₹)', anchor='e')
        self.pnl_tree.heading('Loss', text='Loss (₹)', anchor='e')
        self.pnl_tree.heading('Net P/L', text='Net P/L (₹)', anchor='e')
        self.pnl_tree.heading('Running Total', text='⭐ Accumulated (₹)', anchor='e')

        self.pnl_tree.column('Period', width=200, anchor='w')
        self.pnl_tree.column('Profit', width=130, anchor='e')
        self.pnl_tree.column('Loss', width=130, anchor='e')
        self.pnl_tree.column('Net P/L', width=130, anchor='e')
        self.pnl_tree.column('Running Total', width=160, anchor='e')

        self.pnl_tree.tag_configure('evenrow', background='#f8f9fa')
        self.pnl_tree.tag_configure('oddrow', background='white')

        self.pnl_tree.bind('<MouseWheel>', lambda e: self.pnl_tree.yview_scroll(int(-1 * (e.delta / 120)), "units"))
        self.pnl_tree.pack(fill='both', expand=True)

    def create_equity_summary_section(self, parent: ttk.Frame) -> None:
        """Create equity-wise P/L summary section."""
        ttk.Label(
            parent,
            text="🏷️ EQUITY-WISE P/L SUMMARY",
            font=('Consolas', 13, 'bold')
        ).pack(anchor='w', pady=(0, 10))

        ttk.Label(
            parent,
            text="Net P/L per stock (FIFO + MTF interest applied)",
            font=('Consolas', 9),
            foreground='gray'
        ).pack(anchor='w', pady=(0, 5))

        table_frame = ttk.Frame(parent)
        table_frame.pack(fill='both', expand=True, pady=(0, 10))

        scrollbar = ttk.Scrollbar(table_frame)
        scrollbar.pack(side='right', fill='y')

        columns = ('Equity', 'Closed P/L', 'Open P/L', 'Total')
        self.equity_tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show='headings',
            height=8,
            yscrollcommand=scrollbar.set
        )
        scrollbar.config(command=self.equity_tree.yview)

        self.equity_tree.heading('Equity', text='Equity', anchor='w')
        self.equity_tree.heading('Closed P/L', text='Closed P/L (₹)', anchor='e')
        self.equity_tree.heading('Open P/L', text='Open P/L (₹)', anchor='e')
        self.equity_tree.heading('Total', text='Total (₹)', anchor='e')

        self.equity_tree.column('Equity', width=150, anchor='w')
        self.equity_tree.column('Closed P/L', width=150, anchor='e')
        self.equity_tree.column('Open P/L', width=150, anchor='e')
        self.equity_tree.column('Total', width=150, anchor='e')

        self.equity_tree.tag_configure('evenrow', background='#f8f9fa')
        self.equity_tree.tag_configure('oddrow', background='white')

        self.equity_tree.bind('<MouseWheel>', lambda e: self.equity_tree.yview_scroll(int(-1 * (e.delta / 120)), "units"))
        self.equity_tree.pack(fill='both', expand=True)

    def create_filter_controls(self, parent: ttk.Frame) -> None:
        """Create filter controls section."""
        filter_frame = ttk.LabelFrame(parent, text="📊 Filters", style="Report.TLabelframe")
        filter_frame.pack(fill='x', pady=(0, 20))

        row0 = ttk.Frame(filter_frame)
        row0.pack(fill='x', pady=(0, 10))

        ttk.Label(row0, text="Profiles (combined view):", font=('Arial', 10)).pack(side='left', padx=(0, 5))

        profile_frame = ttk.Frame(row0)
        profile_frame.pack(side='left', padx=5)

        profile_scroll = ttk.Scrollbar(profile_frame, orient='vertical')
        profile_scroll.pack(side='right', fill='y')

        self.profile_listbox = tk.Listbox(
            profile_frame,
            selectmode='extended',
            height=4,
            width=18,
            exportselection=False,
            yscrollcommand=profile_scroll.set
        )
        self.profile_listbox.pack(side='left', fill='y')
        profile_scroll.config(command=self.profile_listbox.yview)

        self.profile_listbox.insert(tk.END, "All")
        self.profile_listbox.selection_set(0)
        Tooltip(self.profile_listbox, "Select profiles to include in Combined Family view")

        def _on_profile_mouse_down(event: tk.Event) -> None:  # type: ignore
            lb = event.widget
            idx = lb.nearest(event.y)
            lb.selection_clear(0, tk.END)
            lb.selection_set(idx)
            lb._anchor = idx

        def _on_profile_drag(event: tk.Event) -> None:  # type: ignore
            lb = event.widget
            try:
                anchor = lb._anchor
            except AttributeError:
                anchor = 0
            idx = lb.nearest(event.y)
            low = min(anchor, idx)
            high = max(anchor, idx)
            lb.selection_clear(0, tk.END)
            lb.selection_set(low, high)

        self.profile_listbox.bind('<Button-1>', _on_profile_mouse_down)
        self.profile_listbox.bind('<B1-Motion>', _on_profile_drag)

        profile_btns = ttk.Frame(row0)
        profile_btns.pack(side='left', padx=(10, 0))

        ttk.Button(
            profile_btns,
            text="Select All",
            command=self.select_all_profiles,
            width=12
        ).pack(anchor='w', pady=(0, 5))

        ttk.Button(
            profile_btns,
            text="Clear",
            command=self.clear_profile_selection,
            width=12
        ).pack(anchor='w')

        ttk.Label(row0, text="(Combined view only)", font=('Arial', 8), foreground='gray').pack(
            side='left', padx=(10, 0)
        )

        row1 = ttk.Frame(filter_frame)
        row1.pack(fill='x', pady=(0, 10))

        ttk.Label(row1, text="Equity (multi-select):", font=('Arial', 10)).pack(side='left', padx=(0, 5))

        equity_frame = ttk.Frame(row1)
        equity_frame.pack(side='left', padx=5)

        equity_scroll = ttk.Scrollbar(equity_frame, orient='vertical')
        equity_scroll.pack(side='right', fill='y')

        self.equity_listbox = tk.Listbox(
            equity_frame,
            selectmode='extended',
            height=4,
            width=18,
            exportselection=False,
            yscrollcommand=equity_scroll.set
        )
        self.equity_listbox.pack(side='left', fill='y')
        equity_scroll.config(command=self.equity_listbox.yview)

        self.equity_listbox.insert(tk.END, "All")
        self.equity_listbox.selection_set(0)
        Tooltip(self.equity_listbox, "Select one or more equities to filter")

        # Enable drag-to-select behavior for usability
        def _on_equity_mouse_down(event: tk.Event) -> None:  # type: ignore
            lb = event.widget
            idx = lb.nearest(event.y)
            lb.selection_clear(0, tk.END)
            lb.selection_set(idx)
            lb._anchor = idx

        def _on_equity_drag(event: tk.Event) -> None:  # type: ignore
            lb = event.widget
            try:
                anchor = lb._anchor
            except AttributeError:
                anchor = 0
            idx = lb.nearest(event.y)
            low = min(anchor, idx)
            high = max(anchor, idx)
            lb.selection_clear(0, tk.END)
            lb.selection_set(low, high)

        self.equity_listbox.bind('<Button-1>', _on_equity_mouse_down)
        self.equity_listbox.bind('<B1-Motion>', _on_equity_drag)

        equity_btns = ttk.Frame(row1)
        equity_btns.pack(side='left', padx=(10, 0))

        ttk.Button(
            equity_btns,
            text="Select All",
            command=self.select_all_equities,
            width=12
        ).pack(anchor='w', pady=(0, 5))

        ttk.Button(
            equity_btns,
            text="Clear",
            command=self.clear_equity_selection,
            width=12
        ).pack(anchor='w')

        row2 = ttk.Frame(filter_frame)
        row2.pack(fill='x', pady=(0, 10))

        ttk.Label(row2, text="From Date:", font=('Arial', 10)).pack(side='left', padx=(0, 5))
        self.from_date_entry = DateEntry(
            row2,
            textvariable=self.from_date_var,
            date_pattern='yyyy-mm-dd',
            width=12
        )
        self.from_date_entry.pack(side='left', padx=5)
        self.from_date_entry.delete(0, 'end')
        Tooltip(self.from_date_entry, "Start date for realized P/L (SELL date)")

        ttk.Label(row2, text="To Date:", font=('Arial', 10)).pack(side='left', padx=(20, 5))
        self.to_date_entry = DateEntry(
            row2,
            textvariable=self.to_date_var,
            date_pattern='yyyy-mm-dd',
            width=12
        )
        self.to_date_entry.pack(side='left', padx=5)
        self.to_date_entry.delete(0, 'end')
        Tooltip(self.to_date_entry, "End date for realized P/L (SELL date)")

        ttk.Label(row2, text="(clear field to disable)", font=('Arial', 8), foreground='gray').pack(
            side='left', padx=(10, 0)
        )

        row3 = ttk.Frame(filter_frame)
        row3.pack(fill='x', pady=(0, 10))

        ttk.Label(row3, text="Type1:", font=('Arial', 10)).pack(side='left', padx=(0, 5))
        self.type1_filter_entry = ttk.Combobox(
            row3,
            textvariable=self.type1_filter_var,
            values=["All", "intraday", "delivery", "mtf", "futures", "options"],
            state='readonly',
            width=12
        )
        self.type1_filter_entry.pack(side='left', padx=(0, 20))
        self.type1_filter_entry.set("All")

        ttk.Label(row3, text="Expiry Month:", font=('Arial', 10)).pack(side='left', padx=(0, 5))
        self.expiry_filter_entry = ttk.Entry(row3, textvariable=self.expiry_month_var, width=12)
        self.expiry_filter_entry.pack(side='left', padx=5)
        ttk.Label(row3, text="(YYYY-MM)", font=('Arial', 8), foreground='gray').pack(
            side='left', padx=(10, 0)
        )

        ttk.Label(row3, text="Preset:", font=('Arial', 10)).pack(side='left', padx=(20, 5))
        self.expiry_preset_entry = ttk.Combobox(
            row3,
            textvariable=self.expiry_preset_var,
            values=["Custom", "Current Month", "Next Month"],
            state='readonly',
            width=14
        )
        self.expiry_preset_entry.pack(side='left', padx=5)
        self.expiry_preset_entry.bind('<<ComboboxSelected>>', lambda _e: self.apply_expiry_preset())
        self.expiry_filter_entry.bind('<KeyRelease>', lambda _e: self.expiry_preset_var.set("Custom"))

        row4 = ttk.Frame(filter_frame)
        row4.pack(fill='x')

        self.open_positions_check = ttk.Checkbutton(
            row4,
            text="Include Open Positions",
            variable=self.show_open_positions_var,
            command=self.toggle_open_positions_display
        )
        self.open_positions_check.pack(side='left', padx=(0, 20))
        Tooltip(self.open_positions_check, "Show holdings that are not fully sold")

        ttk.Button(
            row4,
            text="🔍 Apply Filters",
            command=self.apply_filters,
            width=15
        ).pack(side='left', padx=5)

        ttk.Button(
            row4,
            text="🔄 Clear Filters",
            command=self.clear_filters,
            width=15
        ).pack(side='left', padx=5)

    def create_summary_cards(self, parent: ttk.Frame) -> None:
        """Create summary cards for Total Profit, Total Loss, Net P/L."""
        cards_frame = ttk.Frame(parent)
        cards_frame.pack(fill='x', pady=(0, 20))

        cards_frame.columnconfigure(0, weight=1)
        cards_frame.columnconfigure(1, weight=1)
        cards_frame.columnconfigure(2, weight=1)

        profit_card = ttk.LabelFrame(cards_frame, text="📈 Net Profit", padding="20")
        profit_card.grid(row=0, column=0, padx=15, pady=10, sticky='ew')

        ttk.Label(profit_card, text="Net Profit", font=('Consolas', 9), foreground='gray').pack()

        self.profit_label = ttk.Label(
            profit_card,
            text="₹ +0.00",
            font=('Consolas', 22, 'bold'),
            foreground='#27ae60'
        )
        self.profit_label.pack(pady=(5, 0))

        loss_card = ttk.LabelFrame(cards_frame, text="📉 Net Loss", padding="20")
        loss_card.grid(row=0, column=1, padx=15, pady=10, sticky='ew')

        ttk.Label(loss_card, text="Net Loss", font=('Consolas', 9), foreground='gray').pack()

        self.loss_label = ttk.Label(
            loss_card,
            text="₹ -0.00",
            font=('Consolas', 22, 'bold'),
            foreground='#e74c3c'
        )
        self.loss_label.pack(pady=(5, 0))

        net_card = ttk.LabelFrame(cards_frame, text="🧮 Net P/L", padding="20")
        net_card.grid(row=0, column=2, padx=15, pady=10, sticky='ew')

        ttk.Label(net_card, text="Net P/L", font=('Consolas', 9), foreground='gray').pack()

        value_frame = ttk.Frame(net_card)
        value_frame.pack(pady=(5, 0))

        self.emotion_label = ttk.Label(
            value_frame,
            text="😐",
            font=('Arial', 28)
        )
        self.emotion_label.pack(side='left', padx=(0, 8))

        self.net_label = ttk.Label(
            value_frame,
            text="₹ 0.00",
            font=('Consolas', 22, 'bold')
        )
        self.net_label.pack(side='left')

        # Profile breakdown (used in Combined Family view)
        self.profile_breakdown_frame = ttk.LabelFrame(parent, text="👪 Profile Breakdown", padding="10")
        self.profile_breakdown_frame.pack(fill='x', pady=(10, 10))

        breakdown_header = ttk.Frame(self.profile_breakdown_frame)
        breakdown_header.pack(fill='x', pady=(0, 6))

        ttk.Label(breakdown_header, text="Metric:", font=('Arial', 9)).pack(side='left')
        self.profile_breakdown_mode_entry = ttk.Combobox(
            breakdown_header,
            textvariable=self.profile_breakdown_mode_var,
            values=["Net P/L", "Profit", "Loss"],
            state='readonly',
            width=10
        )
        self.profile_breakdown_mode_entry.pack(side='left', padx=(6, 0))
        self.profile_breakdown_mode_entry.bind('<<ComboboxSelected>>', lambda _e: self._refresh_profile_breakdown())

        self.profile_tree = ttk.Treeview(self.profile_breakdown_frame, columns=('Profile', 'Value'), show='headings', height=4)
        self.profile_tree.heading('Profile', text='Profile')
        self.profile_tree.heading('Value', text='Net P/L (₹)')
        self.profile_tree.column('Profile', width=200, anchor='w')
        self.profile_tree.column('Value', width=120, anchor='e')
        self.profile_tree.pack(fill='x')

    def create_analytics_section(self, parent: ttk.Frame) -> None:
        """Create analytics section for win/loss ratio and holding period."""
        analytics_frame = ttk.LabelFrame(parent, text="📈 Analytics", style="Report.TLabelframe")
        analytics_frame.pack(fill='x', pady=(0, 20))

        for col in range(3):
            analytics_frame.columnconfigure(col, weight=1)

        def add_metric(row: int, col: int, label: str) -> ttk.Label:
            ttk.Label(analytics_frame, text=label, font=('Consolas', 10)).grid(
                row=row, column=col, sticky='w', padx=10, pady=(0, 5)
            )
            value = ttk.Label(
                analytics_frame,
                text="0.00",
                font=('Consolas', 13, 'bold'),
                foreground='#34495e'
            )
            value.grid(row=row + 1, column=col, sticky='w', padx=10)
            return value

        self.win_loss_label = add_metric(0, 0, "Win/Loss Ratio")
        self.win_rate_label = add_metric(0, 1, "Win Rate %")
        self.profit_factor_label = add_metric(0, 2, "Profit Factor")

        self.avg_win_label = add_metric(2, 0, "Avg Win (₹)")
        self.avg_loss_label = add_metric(2, 1, "Avg Loss (₹)")
        self.expectancy_label = add_metric(2, 2, "Expectancy (₹)")

        self.avg_holding_label = add_metric(4, 0, "Avg Holding (days)")
        self.median_holding_label = add_metric(4, 1, "Median Holding (days)")
        self.max_drawdown_label = add_metric(4, 2, "Max Drawdown (₹)")

        self.day_trade_value_label = add_metric(6, 0, "Avg Day Trade Value (₹)")
        self.total_trade_value_label = add_metric(6, 1, "Total Trade Value (₹)")

    def calculate_reports(self) -> None:
        """
        Recalculate all reports from scratch.
        Runs FIFO engine and P/L calculations every time.
        """
        logger.info("=" * 60)
        logger.info("Starting P/L report calculation")
        logger.info("=" * 60)

        self.update_status("⏳ Calculating P/L reports...")

        try:
            import config as _config

            original_profile = _config.CURRENT_PROFILE_ID

            combined_filtered_pnl_results: list = []
            combined_trades_by_id: dict = {}
            profile_pnls: dict[str, int] = {}
            profile_breakdown: dict[str, dict[str, int]] = {}
            combined_matches: list = []

            try:
                conn = sqlite3.connect(str(config.DB_PATH))
                cur = conn.cursor()
                cur.execute("SELECT profile_name FROM profiles WHERE is_active = 1 ORDER BY profile_name")
                profile_names = [row[0] for row in cur.fetchall()]
                conn.close()
                self._set_profile_listbox_values(["All"] + profile_names)
            except Exception:
                pass

            # If Combined Family view selected (0), compute per-profile and aggregate
            if _config.CURRENT_PROFILE_ID == 0:
                # Load active profiles
                conn = sqlite3.connect(str(config.DB_PATH))
                cur = conn.cursor()
                cur.execute("SELECT id, profile_name FROM profiles WHERE is_active = 1 ORDER BY profile_name")
                profiles = cur.fetchall()
                conn.close()

                if not profiles:
                    messagebox.showinfo("No Profiles", "No active profiles found in database.")
                    self.update_status("⚠️ No profiles to analyze")
                    self.reset_displays()
                    return

                selected_profiles = self._get_selected_profiles()
                if selected_profiles:
                    profiles = [
                        (pid, pname)
                        for pid, pname in profiles
                        if pname in selected_profiles
                    ]

                if not profiles:
                    messagebox.showinfo("No Profiles", "No matching profiles selected.")
                    self.update_status("⚠️ No profiles selected")
                    self.reset_displays()
                    return

                for pid, pname in profiles:
                    try:
                        _config.CURRENT_PROFILE_ID = pid
                        trades = fetch_active_trades()
                        logger.info(f"Fetched {len(trades)} active trades for profile {pname}")
                        if not trades:
                            profile_pnls[pname] = 0
                            profile_breakdown[pname] = {"profit": 0, "loss": 0, "net": 0}
                            continue

                        matches = match_fifo(trades)
                        if not matches:
                            profile_pnls[pname] = 0
                            profile_breakdown[pname] = {"profit": 0, "loss": 0, "net": 0}
                            continue

                        trades_by_id = build_trades_by_id(trades)
                        pnl_results = calculate_match_pnl(matches, trades_by_id)
                        pnl_results = apply_mtf_interest(pnl_results, trades_by_id)
                        combined_matches.extend(matches)

                        # Apply UI filters per-profile (date/equity/type1/expiry)
                        filtered_pnl_results = pnl_results

                        from_date = (self.from_date_entry.get().strip() if hasattr(self, 'from_date_entry') else self.from_date_var.get().strip()) or None
                        to_date = (self.to_date_entry.get().strip() if hasattr(self, 'to_date_entry') else self.to_date_var.get().strip()) or None
                        if from_date or to_date:
                            filtered_pnl_results = filter_matches_by_date_range(filtered_pnl_results, trades_by_id, from_date, to_date)

                        selected_equities = self._get_selected_equities()
                        if selected_equities:
                            filtered_pnl_results = [
                                pnl for pnl in filtered_pnl_results
                                if trades_by_id[pnl['sell_id']]['equity'] in selected_equities
                            ]

                        type1_filter = self.type1_filter_var.get().strip().lower()
                        if type1_filter and type1_filter != "all":
                            filtered_pnl_results = [
                                pnl for pnl in filtered_pnl_results
                                if (trades_by_id[pnl['sell_id']].get('type1') or "delivery") == type1_filter
                            ]

                        expiry_month = self.expiry_filter_entry.get().strip() if hasattr(self, 'expiry_filter_entry') else self.expiry_month_var.get().strip()
                        if expiry_month:
                            filtered_pnl_results = [
                                pnl for pnl in filtered_pnl_results
                                if (trades_by_id[pnl['sell_id']].get('expiry') or "").startswith(expiry_month)
                            ]

                        # Sum net_pnl for this profile
                        profile_profit = sum(pnl['net_pnl'] for pnl in filtered_pnl_results if pnl['net_pnl'] > 0)
                        profile_loss = sum(pnl['net_pnl'] for pnl in filtered_pnl_results if pnl['net_pnl'] < 0)
                        profile_total = profile_profit + profile_loss
                        profile_pnls[pname] = profile_total
                        profile_breakdown[pname] = {
                            "profit": profile_profit,
                            "loss": profile_loss,
                            "net": profile_total
                        }

                        # Accumulate for combined reporting
                        combined_filtered_pnl_results.extend(filtered_pnl_results)
                        combined_trades_by_id.update(trades_by_id)

                    except Exception as e:
                        logger.warning(f"Failed to compute P/L for profile {pname}: {e}")

                # Restore original profile selection
                _config.CURRENT_PROFILE_ID = original_profile

                # Use combined results for downstream aggregations
                filtered_pnl_results = combined_filtered_pnl_results
                trades_by_id = combined_trades_by_id
                matches = combined_matches

                logger.info(f"Aggregated combined P/L across {len(profile_pnls)} profiles")

            else:
                trades = fetch_active_trades()
                logger.info(f"Fetched {len(trades)} active trades")

                if not trades:
                    messagebox.showinfo("No Data", "No active trades found in database.")
                    self.update_status("⚠️ No trades to analyze")
                    self.reset_displays()
                    return

                matches = match_fifo(trades)
                logger.info(f"Generated {len(matches) if matches else 0} FIFO matches")

                if not matches:
                    messagebox.showinfo("No Matches", "No SELL trades found. P/L can only be calculated after selling.")
                    self.update_status("⚠️ No realized P/L yet")
                    self.reset_displays()
                    return

                trades_by_id = build_trades_by_id(trades)
                pnl_results = calculate_match_pnl(matches, trades_by_id)
                pnl_results = apply_mtf_interest(pnl_results, trades_by_id)
                logger.info(f"Calculated P/L for {len(pnl_results)} matches")

                filtered_pnl_results = pnl_results

            from_date = (self.from_date_entry.get().strip() if hasattr(self, 'from_date_entry') else self.from_date_var.get().strip()) or None
            to_date = (self.to_date_entry.get().strip() if hasattr(self, 'to_date_entry') else self.to_date_var.get().strip()) or None
            if from_date or to_date:
                filtered_pnl_results = filter_matches_by_date_range(filtered_pnl_results, trades_by_id, from_date, to_date)
                logger.info(f"Date filter applied: {len(filtered_pnl_results)} matches in range")

            selected_equities = self._get_selected_equities()
            if selected_equities:
                filtered_pnl_results = [
                    pnl for pnl in filtered_pnl_results
                    if trades_by_id[pnl['sell_id']]['equity'] in selected_equities
                ]
                logger.info(f"Equity filter {selected_equities} applied: {len(filtered_pnl_results)} matches")

            type1_filter = self.type1_filter_var.get().strip().lower()
            if type1_filter and type1_filter != "all":
                filtered_pnl_results = [
                    pnl for pnl in filtered_pnl_results
                    if (trades_by_id[pnl['sell_id']].get('type1') or "delivery") == type1_filter
                ]
                logger.info(f"Type1 filter {type1_filter} applied: {len(filtered_pnl_results)} matches")

            expiry_month = self.expiry_filter_entry.get().strip() if hasattr(self, 'expiry_filter_entry') else self.expiry_month_var.get().strip()
            if expiry_month:
                filtered_pnl_results = [
                    pnl for pnl in filtered_pnl_results
                    if (trades_by_id[pnl['sell_id']].get('expiry') or "").startswith(expiry_month)
                ]
                logger.info(f"Expiry month filter {expiry_month} applied: {len(filtered_pnl_results)} matches")

            self.audit_matches = list(filtered_pnl_results)
            self.audit_trades_by_id = trades_by_id

            # If combined view, update profile breakdown UI
            try:
                import config as _config
                if _config.CURRENT_PROFILE_ID == 0:
                    self.profile_breakdown_data = profile_breakdown
                else:
                    self.profile_breakdown_data = {}
                self._refresh_profile_breakdown()
            except Exception:
                pass

            self.open_positions = calculate_open_positions(matches, trades_by_id)
            self.filtered_open_positions = self._filter_open_positions(self.open_positions)
            equities = ["All"] + get_unique_equities(trades_by_id)
            self._set_equity_listbox_values(equities)

            equity_pnl_totals = aggregate_pnl_by_equity(filtered_pnl_results, trades_by_id, pnl_field="net_pnl")
            self.equity_pnl = self._convert_to_pnl_breakdown(equity_pnl_totals)

            daily_pnl_totals = aggregate_pnl_by_date(filtered_pnl_results, trades_by_id, pnl_field="net_pnl")
            weekly_pnl_totals = aggregate_pnl_by_week(daily_pnl_totals)
            monthly_pnl_totals = aggregate_pnl_by_month(daily_pnl_totals)
            yearly_pnl_totals = aggregate_pnl_by_year(monthly_pnl_totals)

            self.daily_pnl = self._convert_to_pnl_breakdown(daily_pnl_totals)
            self.weekly_pnl = self._convert_to_pnl_breakdown(weekly_pnl_totals)
            self.monthly_pnl = self._convert_to_pnl_breakdown(monthly_pnl_totals)
            self.yearly_pnl = self._convert_to_pnl_breakdown(yearly_pnl_totals)

            self.total_profit = sum(pnl['net_pnl'] for pnl in filtered_pnl_results if pnl['net_pnl'] > 0)
            self.total_loss = sum(pnl['net_pnl'] for pnl in filtered_pnl_results if pnl['net_pnl'] < 0)
            self.net_pnl = self.total_profit + self.total_loss

            sell_totals = aggregate_pnl_by_sell(filtered_pnl_results, pnl_field="net_pnl")
            wins = sum(1 for pnl in sell_totals.values() if pnl > 0)
            losses = sum(1 for pnl in sell_totals.values() if pnl < 0)
            total_trades = len(sell_totals)
            win_rate = (wins / total_trades * 100) if total_trades > 0 else 0.0

            win_values = [pnl for pnl in sell_totals.values() if pnl > 0]
            loss_values = [pnl for pnl in sell_totals.values() if pnl < 0]
            avg_win = float(sum(win_values) / len(win_values)) if win_values else 0
            avg_loss = float(sum(loss_values) / len(loss_values)) if loss_values else 0

            if losses == 0:
                win_loss_ratio = "∞" if wins > 0 else "0.00"
            else:
                win_loss_ratio = f"{wins / losses:.2f}"

            total_profit_val = sum(win_values)
            total_loss_abs = abs(sum(loss_values))
            if total_loss_abs == 0:
                profit_factor = "∞" if total_profit_val > 0 else "0.00"
            else:
                profit_factor = f"{total_profit_val / total_loss_abs:.2f}"

            loss_rate = 1 - (wins / total_trades) if total_trades > 0 else 0.0
            expectancy = int((avg_win * (win_rate / 100)) + (avg_loss * loss_rate))

            total_qty = 0
            total_days = 0
            holding_buckets: list[tuple[int, int]] = []
            for match in filtered_pnl_results:
                buy_date = trades_by_id[match['buy_id']]['trade_date']
                sell_date = trades_by_id[match['sell_id']]['trade_date']
                days = (datetime.strptime(sell_date, '%Y-%m-%d') - datetime.strptime(buy_date, '%Y-%m-%d')).days
                qty = match['matched_quantity']
                total_days += days * qty
                total_qty += qty
                holding_buckets.append((days, qty))
            avg_holding_days = (total_days / total_qty) if total_qty > 0 else 0.0
            median_holding_days = self._weighted_median(holding_buckets) if holding_buckets else 0.0

            max_drawdown = self._calculate_max_drawdown(daily_pnl_totals)

            # Derived trade value metrics (based on matched trades)
            trade_value_by_date = aggregate_trade_value_by_date(filtered_pnl_results, trades_by_id)
            self.analytics['trade_value_by_date'] = trade_value_by_date

            self.analytics = {
                'win_loss_ratio': win_loss_ratio,
                'win_rate': win_rate,
                'avg_win': avg_win,
                'avg_loss': avg_loss,
                'profit_factor': profit_factor,
                'expectancy': expectancy,
                'avg_holding_days': avg_holding_days,
                'median_holding_days': median_holding_days,
                'max_drawdown': max_drawdown
            }

            self.hide_warning_banner()
            self.update_displays()

            self.update_status(f"✅ FIFO validated | Net P/L = ₹{self.net_pnl / 100:.2f}")
            logger.info("=" * 60)
            logger.info("P/L report calculation completed successfully")
            logger.info("=" * 60)

        except Exception as e:
            logger.error(f"❌ Failed to calculate reports: {str(e)}", exc_info=True)
            messagebox.showerror("Error", f"Failed to calculate reports:\n\n{str(e)}")
            self.update_status("❌ Error calculating reports")
            self.reset_displays()

    def update_displays(self) -> None:
        """Update all display elements with calculated values."""
        self.profit_label.config(text=format_money(self.total_profit))

        if self.total_loss < 0:
            loss_formatted = format_money_abs(self.total_loss)
            self.loss_label.config(text=f"₹ -{loss_formatted.replace('₹', '').strip()}")
        else:
            self.loss_label.config(text="₹ 0.00")

        self.net_label.config(text=format_money(self.net_pnl))

        if self.net_pnl > 0:
            self.emotion_label.config(text="🙂")
            self.net_label.config(foreground='#27ae60')
        elif self.net_pnl < 0:
            self.emotion_label.config(text="😢")
            self.net_label.config(foreground='#e74c3c')
        else:
            self.emotion_label.config(text="😐")
            self.net_label.config(foreground='#34495e')

        # Derived trade value metrics
        trade_value_by_date = self.analytics.get('trade_value_by_date') or {}
        total_trade_value = sum(trade_value_by_date.values()) if trade_value_by_date else 0
        avg_day_trade_value = 0
        if trade_value_by_date:
            avg_day_trade_value = int(total_trade_value / max(len(trade_value_by_date), 1))
        self.day_trade_value_label.config(text=format_money(avg_day_trade_value))
        self.total_trade_value_label.config(text=format_money(total_trade_value))

        self.update_period_display()

        if hasattr(self, 'win_loss_label'):
            self.win_loss_label.config(text=str(self.analytics.get('win_loss_ratio', '0.00')))
        if hasattr(self, 'win_rate_label'):
            win_rate = self.analytics.get('win_rate', 0.0)
            self.win_rate_label.config(text=f"{win_rate:.1f}%")
        if hasattr(self, 'profit_factor_label'):
            self.profit_factor_label.config(text=str(self.analytics.get('profit_factor', '0.00')))
        if hasattr(self, 'avg_win_label'):
            self.avg_win_label.config(text=format_money(self.analytics.get('avg_win', 0)))
        if hasattr(self, 'avg_loss_label'):
            self.avg_loss_label.config(text=format_money(self.analytics.get('avg_loss', 0)))
        if hasattr(self, 'expectancy_label'):
            self.expectancy_label.config(text=format_money(self.analytics.get('expectancy', 0)))
        if hasattr(self, 'avg_holding_label'):
            avg_days = self.analytics.get('avg_holding_days', 0.0)
            self.avg_holding_label.config(text=f"{avg_days:.1f}")
        if hasattr(self, 'median_holding_label'):
            median_days = self.analytics.get('median_holding_days', 0.0)
            self.median_holding_label.config(text=f"{median_days:.1f}")
        if hasattr(self, 'max_drawdown_label'):
            self.max_drawdown_label.config(text=format_money(self.analytics.get('max_drawdown', 0)))

        self.update_equity_summary()

        if self.show_open_positions_var.get():
            self.update_open_positions_table()

    def update_period_display(self) -> None:
        """Update the P/L table based on selected period."""
        period = self.period_var.get()

        title_map = {
            "Daily": "📅 DAILY P/L",
            "Weekly": "🗓️ WEEKLY P/L",
            "Monthly": "🗓️ MONTHLY P/L",
            "Yearly": "📆 YEARLY P/L"
        }
        self.pnl_title_label.config(text=title_map.get(period, "P/L BREAKDOWN"))

        for item in self.pnl_tree.get_children():
            self.pnl_tree.delete(item)

        period_map = {
            "Daily": (self.daily_pnl, "daily"),
            "Weekly": (self.weekly_pnl, "weekly"),
            "Monthly": (self.monthly_pnl, "monthly"),
            "Yearly": (self.yearly_pnl, "yearly")
        }

        if period not in period_map:
            return

        data_dict, mode = period_map[period]

        if not data_dict:
            _ = self.pnl_tree.insert('', 'end', values=(
                "No realized P/L in this period",
                "",
                "",
                "(no SELL trades)",
                ""
            ), tags=('empty',))
            self.pnl_tree.tag_configure('empty', foreground='gray')
            return

        rows = self.build_report_rows(data_dict, mode)

        for idx, row in enumerate(reversed(rows)):
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'

            _ = self.pnl_tree.insert('', 'end', values=(
                row['label'],
                format_money(row['profit']) if row['profit'] != 0 else "₹ 0.00",
                format_money(row['loss']) if row['loss'] != 0 else "₹ 0.00",
                format_money(row['net']),
                f"→ {format_money(row['accumulated'])}"
            ), tags=(tag,))

    def show_warning_banner(self) -> None:
        """Show warning banner at top of Reports tab."""
        if self.has_validation_errors and hasattr(self, 'warning_frame'):
            self.warning_label.config(text=f"⚠️ VALIDATION ERROR: {self.validation_message}")
            self.warning_frame.pack(fill='x', pady=(0, 10), before=self.warning_frame.master.winfo_children()[1])

    def hide_warning_banner(self) -> None:
        """Hide warning banner."""
        if hasattr(self, 'warning_frame'):
            self.warning_frame.pack_forget()

    def reset_displays(self) -> None:
        """Reset all displays to zero/empty state."""
        self.total_profit = 0
        self.total_loss = 0
        self.net_pnl = 0
        self.daily_pnl = {}
        self.weekly_pnl = {}
        self.monthly_pnl = {}
        self.yearly_pnl = {}
        self.analytics = {}
        self.open_positions = []
        self.filtered_open_positions = []
        self.equity_pnl = {}
        self.audit_matches = []
        self.audit_trades_by_id = {}
        self.profile_breakdown_data = {}

        self.profit_label.config(text="₹ +0.00")
        self.loss_label.config(text="₹ -0.00")
        self.net_label.config(text="₹ 0.00", foreground='#34495e')
        self.emotion_label.config(text="😐")

        if hasattr(self, 'win_loss_label'):
            self.win_loss_label.config(text="0.00")
        if hasattr(self, 'win_rate_label'):
            self.win_rate_label.config(text="0.0%")
        if hasattr(self, 'profit_factor_label'):
            self.profit_factor_label.config(text="0.00")
        if hasattr(self, 'avg_win_label'):
            self.avg_win_label.config(text="₹ 0.00")
        if hasattr(self, 'avg_loss_label'):
            self.avg_loss_label.config(text="₹ 0.00")
        if hasattr(self, 'expectancy_label'):
            self.expectancy_label.config(text="₹ 0.00")
        if hasattr(self, 'avg_holding_label'):
            self.avg_holding_label.config(text="0.0")
        if hasattr(self, 'median_holding_label'):
            self.median_holding_label.config(text="0.0")
        if hasattr(self, 'max_drawdown_label'):
            self.max_drawdown_label.config(text="₹ 0.00")

        for item in self.pnl_tree.get_children():
            self.pnl_tree.delete(item)

        if hasattr(self, 'profile_tree'):
            self.profile_tree.delete(*self.profile_tree.get_children())

    def print_report(self) -> None:
        """Generate a print-friendly HTML report."""
        logger.info("Generating print report")

        try:
            import webbrowser

            config.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = config.EXPORTS_DIR / f"report_{timestamp}.html"

            period_type = self.period_var.get()

            html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>P/L Report - {datetime.now().strftime("%d %b %Y %I:%M %p")}</title>
    <style>
        @media print {{
            @page {{ margin: 1cm; }}
            body {{ margin: 0; }}
        }}

        body {{
            font-family: Consolas, 'Courier New', monospace;
            max-width: 1000px;
            margin: 20px auto;
            padding: 20px;
            background: white;
        }}

        h1 {{
            text-align: center;
            color: #2c3e50;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }}

        .timestamp {{
            text-align: center;
            color: #7f8c8d;
            margin-bottom: 30px;
        }}

        .summary {{
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 20px;
            margin-bottom: 30px;
        }}

        .card {{
            border: 2px solid #ecf0f1;
            border-radius: 8px;
            padding: 15px;
            text-align: center;
        }}

        .card-title {{
            font-size: 12px;
            color: #7f8c8d;
            margin-bottom: 10px;
        }}

        .card-value {{
            font-size: 24px;
            font-weight: bold;
        }}

        .profit {{ color: #27ae60; }}
        .loss {{ color: #e74c3c; }}

        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}

        th {{
            background: #34495e;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }}

        td {{
            padding: 10px;
            border-bottom: 1px solid #ecf0f1;
        }}

        tr:hover {{
            background: #f8f9fa;
        }}

        .footer {{
            margin-top: 40px;
            text-align: center;
            color: #7f8c8d;
            font-size: 12px;
            border-top: 1px solid #ecf0f1;
            padding-top: 20px;
        }}
    </style>
</head>
<body>
    <h1>📊 PROFIT/LOSS REPORT</h1>
    <div class="timestamp">Generated on {datetime.now().strftime("%d %b %Y at %I:%M %p")}</div>

    <div class="summary">
        <div class="card">
            <div class="card-title">Total Profit</div>
            <div class="card-value profit">{format_money(self.total_profit)}</div>
        </div>
        <div class="card">
            <div class="card-title">Total Loss</div>
            <div class="card-value loss">{format_money_abs(self.total_loss)}</div>
        </div>
        <div class="card">
            <div class="card-title">Net P/L</div>
            <div class="card-value {'profit' if self.net_pnl >= 0 else 'loss'}">{format_money(self.net_pnl)}</div>
        </div>
    </div>

    <h2>{period_type} P/L Breakdown</h2>
    <table>
        <thead>
            <tr>
                <th>Period</th>
                <th>Profit</th>
                <th>Loss</th>
                <th>Net P/L</th>
                <th>Running Total</th>
            </tr>
        </thead>
        <tbody>
"""

            for item_id in self.pnl_tree.get_children():
                values = self.pnl_tree.item(item_id, 'values')
                if len(values) >= 5:
                    period_name, profit_str, loss_str, net_str, running_str = values
                    row_class = 'profit' if '₹' in net_str and '-' not in net_str else 'loss' if '-' in net_str else ''

                    html_content += f"""
            <tr>
                <td>{period_name}</td>
                <td class="profit">{profit_str}</td>
                <td class="loss">{loss_str}</td>
                <td class="{row_class}">{net_str}</td>
                <td class="{'profit' if '-' not in running_str else 'loss'}">{running_str}</td>
            </tr>
"""

            html_content += """
        </tbody>
    </table>

    <div class="footer">
        <p>Trader Ledger - FIFO-based P/L Calculation System</p>
        <p>This report was automatically generated. Please verify all figures.</p>
    </div>
</body>
</html>
"""

            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_content)

            logger.info(f"✅ Report saved: {filepath}")
            webbrowser.open(f'file:///{Path(filepath).absolute()}')

            messagebox.showinfo(
                "Report Generated",
                f"Print-friendly report created:\n{filepath}\n\nOpening in your default browser..."
            )

        except Exception as e:
            logger.error(f"Print report failed: {str(e)}", exc_info=True)
            messagebox.showerror("Print Failed", f"Failed to generate report:\n{str(e)}")

    def _collect_report_export_data(self) -> dict:
        """Collect current filtered report data for exports."""
        selected_equities = self._get_selected_equities()
        equity_label = ", ".join(selected_equities) if selected_equities else "All"

        data = {
            'filters': {
                'equities': equity_label,
                'from_date': self.from_date_entry.get().strip() if hasattr(self, 'from_date_entry') else self.from_date_var.get().strip(),
                'to_date': self.to_date_entry.get().strip() if hasattr(self, 'to_date_entry') else self.to_date_var.get().strip(),
                'type1': self.type1_filter_var.get(),
                'expiry_month': self.expiry_filter_entry.get().strip() if hasattr(self, 'expiry_filter_entry') else self.expiry_month_var.get().strip(),
                'include_open_positions': self.show_open_positions_var.get()
            },
            'summary': {
                'total_profit': format_money(self.total_profit),
                'total_loss': format_money_abs(self.total_loss),
                'net_pnl': format_money(self.net_pnl),
                'win_loss_ratio': self.analytics.get('win_loss_ratio', '0.00'),
                'win_rate': f"{self.analytics.get('win_rate', 0.0):.1f}%",
                'profit_factor': self.analytics.get('profit_factor', '0.00'),
                'avg_win': format_money(self.analytics.get('avg_win', 0)),
                'avg_loss': format_money(self.analytics.get('avg_loss', 0)),
                'expectancy': format_money(self.analytics.get('expectancy', 0)),
                'avg_holding_days': f"{self.analytics.get('avg_holding_days', 0.0):.1f}",
                'median_holding_days': f"{self.analytics.get('median_holding_days', 0.0):.1f}",
                'max_drawdown': format_money(self.analytics.get('max_drawdown', 0))
            },
            'period': {
                'type': self.period_var.get(),
                'rows': []
            },
            'equity_summary': [],
            'open_positions': []
        }

        for item_id in self.pnl_tree.get_children():
            values = self.pnl_tree.item(item_id, 'values')
            if len(values) >= 5:
                data['period']['rows'].append(values)

        if hasattr(self, 'equity_tree'):
            for item_id in self.equity_tree.get_children():
                values = self.equity_tree.item(item_id, 'values')
                if len(values) >= 4:
                    data['equity_summary'].append(values)

        if hasattr(self, 'open_tree') and self.show_open_positions_var.get():
            for item_id in self.open_tree.get_children():
                values = self.open_tree.item(item_id, 'values')
                if len(values) >= 10:
                    data['open_positions'].append(values)

        return data

    def _get_trade_ts_map(self, trade_ids: set[int]) -> dict[int, str]:
        """Fetch trade timestamps for audit export (fallback to 09:15:00)."""
        if not trade_ids:
            return {}

        ids = sorted(trade_ids)
        placeholders = ",".join("?" for _ in ids)
        conn = sqlite3.connect(str(config.DB_PATH))
        cursor = conn.cursor()
        cursor.execute(
            f"SELECT id, trade_date, trade_ts FROM trade_events WHERE id IN ({placeholders})",
            ids
        )
        rows = cursor.fetchall()
        conn.close()

        trade_ts_map: dict[int, str] = {}
        for trade_id, trade_date, trade_ts in rows:
            if trade_ts:
                trade_ts_map[trade_id] = trade_ts
            elif trade_date:
                trade_ts_map[trade_id] = f"{trade_date} 09:15:00"
            else:
                trade_ts_map[trade_id] = ""

        return trade_ts_map

    def _build_remainder_flag_by_match_index(self) -> dict[int, str]:
        """Mark rows where deterministic allocation remainder was applied."""
        flags: dict[int, list[str]] = {}
        if not self.audit_matches:
            return {}

        buy_indices: dict[int, list[int]] = {}
        sell_indices: dict[int, list[int]] = {}
        for idx, match in enumerate(self.audit_matches):
            buy_indices.setdefault(match['buy_id'], []).append(idx)
            sell_indices.setdefault(match['sell_id'], []).append(idx)

        def has_remainder(total_amount: int, trade_qty: int, qtys: list[int]) -> bool:
            matched_total = sum(qtys)
            if trade_qty <= 0 or matched_total <= 0:
                return False
            proportional_total = round_divide(total_amount * matched_total, trade_qty)
            base_alloc_sum = sum((proportional_total * qty) // matched_total for qty in qtys)
            return (proportional_total - base_alloc_sum) != 0

        for buy_id, idxs in buy_indices.items():
            buy_trade = self.audit_trades_by_id.get(buy_id)
            if not buy_trade:
                continue
            qtys = [self.audit_matches[i]['matched_quantity'] for i in idxs]

            if has_remainder(int(buy_trade.get('brokerage', 0) or 0), int(buy_trade.get('quantity', 0) or 0), qtys):
                flags.setdefault(idxs[-1], []).append('BUY_BRK')

            if (buy_trade.get('type1') or '').lower() == 'mtf':
                if has_remainder(int(buy_trade.get('mtf_amount', 0) or 0), int(buy_trade.get('quantity', 0) or 0), qtys):
                    flags.setdefault(idxs[-1], []).append('BUY_MTF')

        for sell_id, idxs in sell_indices.items():
            sell_trade = self.audit_trades_by_id.get(sell_id)
            if not sell_trade:
                continue
            qtys = [self.audit_matches[i]['matched_quantity'] for i in idxs]
            if has_remainder(int(sell_trade.get('brokerage', 0) or 0), int(sell_trade.get('quantity', 0) or 0), qtys):
                flags.setdefault(idxs[-1], []).append('SELL_BRK')

        return {idx: '|'.join(parts) for idx, parts in flags.items()}

    def export_audit_csv(self) -> None:
        """Export match-level audit details to CSV."""
        try:
            if not self.audit_matches:
                messagebox.showwarning(
                    "No Data",
                    "No match-level data to export. Click Recalculate first."
                )
                return

            config.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = config.EXPORTS_DIR / f"audit_{timestamp}.csv"

            trade_ids = {m['buy_id'] for m in self.audit_matches} | {m['sell_id'] for m in self.audit_matches}
            trade_ts_map = self._get_trade_ts_map(trade_ids)
            remainder_flags = self._build_remainder_flag_by_match_index()

            def rupees(paise: int, absolute: bool = False) -> str:
                value = abs(paise) if absolute else paise
                return f"{value / 100:.2f}"

            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([
                    "SellID", "BuyID", "Equity", "Type1", "Type2", "Strike", "Expiry",
                    "MatchedQty", "BuyDate", "BuyTS", "SellDate", "SellTS",
                    "BuyBrokerageAutoPaise", "BuyBrokerageOverridePaise", "BuyBrokerageEffectivePaise",
                    "SellBrokerageAutoPaise", "SellBrokerageOverridePaise", "SellBrokerageEffectivePaise",
                    "BuyPrice", "SellPrice", "BuyCost", "SellValue",
                    "BuyBrokerageAlloc", "SellBrokerageAlloc",
                    "MatchedBuyTotal", "MatchedSellTotal", "GrossPnL",
                    "MatchedMtfAmount", "HoldingDays", "MtfInterest", "NetPnL",
                    "AllocationRemainderApplied",
                    "BuyPricePaise", "SellPricePaise", "BuyCostPaise", "SellValuePaise",
                    "BuyBrokerageAllocPaise", "SellBrokerageAllocPaise",
                    "MatchedBuyTotalPaise", "MatchedSellTotalPaise", "GrossPnLPaise",
                    "MatchedMtfAmountPaise", "MtfInterestPaise", "NetPnLPaise"
                ])

                for idx, match in enumerate(self.audit_matches):
                    buy = self.audit_trades_by_id.get(match['buy_id'])
                    sell = self.audit_trades_by_id.get(match['sell_id'])
                    if not buy or not sell:
                        continue

                    buy_date = buy.get('trade_date') or ""
                    sell_date = sell.get('trade_date') or ""
                    buy_ts = trade_ts_map.get(match['buy_id']) or (f"{buy_date} 09:15:00" if buy_date else "")
                    sell_ts = trade_ts_map.get(match['sell_id']) or (f"{sell_date} 09:15:00" if sell_date else "")

                    gross_pnl = match.get('gross_pnl', match.get('realized_pnl', 0))
                    net_pnl = match.get('net_pnl', gross_pnl)
                    buy_brokerage_override = buy.get('brokerage_override')
                    sell_brokerage_override = sell.get('brokerage_override')

                    writer.writerow([
                        match['sell_id'],
                        match['buy_id'],
                        sell.get('equity', ''),
                        sell.get('type1', ''),
                        sell.get('type2', '') or '',
                        "" if sell.get('strike') is None else str(sell.get('strike')),
                        sell.get('expiry', '') or '',
                        match['matched_quantity'],
                        buy_date,
                        buy_ts,
                        sell_date,
                        sell_ts,
                        int(buy.get('brokerage_auto', 0) or 0),
                        "" if buy_brokerage_override is None else int(buy_brokerage_override),
                        int(buy.get('brokerage', 0) or 0),
                        int(sell.get('brokerage_auto', 0) or 0),
                        "" if sell_brokerage_override is None else int(sell_brokerage_override),
                        int(sell.get('brokerage', 0) or 0),
                        rupees(int(buy.get('price', 0)), absolute=True),
                        rupees(int(sell.get('price', 0)), absolute=True),
                        rupees(match['buy_cost'], absolute=True),
                        rupees(match['sell_value'], absolute=True),
                        rupees(match['buy_brokerage_alloc'], absolute=True),
                        rupees(match['sell_brokerage_alloc'], absolute=True),
                        rupees(match.get('matched_buy_total', match['buy_cost'] + match['buy_brokerage_alloc']), absolute=True),
                        rupees(match.get('matched_sell_total', match['sell_value'] - match['sell_brokerage_alloc']), absolute=True),
                        rupees(gross_pnl),
                        rupees(int(match.get('matched_mtf_amount', 0)), absolute=True),
                        match.get('holding_days', 0),
                        rupees(int(match.get('mtf_interest', 0)), absolute=True),
                        rupees(net_pnl),
                        remainder_flags.get(idx, "NONE"),
                        int(buy.get('price', 0) or 0),
                        int(sell.get('price', 0) or 0),
                        int(match['buy_cost']),
                        int(match['sell_value']),
                        int(match['buy_brokerage_alloc']),
                        int(match['sell_brokerage_alloc']),
                        int(match.get('matched_buy_total', match['buy_cost'] + match['buy_brokerage_alloc'])),
                        int(match.get('matched_sell_total', match['sell_value'] - match['sell_brokerage_alloc'])),
                        int(gross_pnl),
                        int(match.get('matched_mtf_amount', 0) or 0),
                        int(match.get('mtf_interest', 0) or 0),
                        int(net_pnl)
                    ])

            messagebox.showinfo("Export Complete", f"Audit CSV saved to:\n{filepath}")

        except Exception as e:
            logger.error(f"Audit CSV export failed: {str(e)}", exc_info=True)
            messagebox.showerror("Export Failed", f"Failed to export audit CSV:\n{str(e)}")

    def export_report_csv(self) -> None:
        """Export current filtered report to CSV."""
        try:
            config.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = config.EXPORTS_DIR / f"report_{timestamp}.csv"

            data = self._collect_report_export_data()

            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)

                writer.writerow(["Filters"])
                writer.writerow(["Equities", data['filters']['equities']])
                writer.writerow(["From Date", data['filters']['from_date'] or "-"])
                writer.writerow(["To Date", data['filters']['to_date'] or "-"])
                writer.writerow(["Type1", data['filters']['type1'] or "All"])
                writer.writerow(["Expiry Month", data['filters']['expiry_month'] or "-"])
                writer.writerow(["Include Open Positions", str(data['filters']['include_open_positions'])])
                writer.writerow([])

                writer.writerow(["Summary"])
                writer.writerow(["Total Profit", data['summary']['total_profit']])
                writer.writerow(["Total Loss", data['summary']['total_loss']])
                writer.writerow(["Net P/L", data['summary']['net_pnl']])
                writer.writerow(["Win/Loss Ratio", data['summary']['win_loss_ratio']])
                writer.writerow(["Win Rate", data['summary']['win_rate']])
                writer.writerow(["Profit Factor", data['summary']['profit_factor']])
                writer.writerow(["Avg Win", data['summary']['avg_win']])
                writer.writerow(["Avg Loss", data['summary']['avg_loss']])
                writer.writerow(["Expectancy", data['summary']['expectancy']])
                writer.writerow(["Avg Holding Period (days)", data['summary']['avg_holding_days']])
                writer.writerow(["Median Holding Period (days)", data['summary']['median_holding_days']])
                writer.writerow(["Max Drawdown", data['summary']['max_drawdown']])
                writer.writerow([])

                writer.writerow([f"{data['period']['type']} P/L"])
                writer.writerow(["Period", "Profit", "Loss", "Net P/L", "Running Total"])
                for row in data['period']['rows']:
                    writer.writerow(row)
                writer.writerow([])

                writer.writerow(["Equity-wise Summary"])
                writer.writerow(["Equity", "Closed P/L", "Open P/L", "Total"])
                for row in data['equity_summary']:
                    writer.writerow(row)
                writer.writerow([])

                if data['open_positions']:
                    writer.writerow(["Open Positions"])
                    writer.writerow([
                        "Equity", "Type1", "Type2", "Strike", "Expiry",
                        "Holding Days", "Status", "Qty", "Avg Price", "Unrealized P/L"
                    ])
                    for row in data['open_positions']:
                        writer.writerow(row)

            messagebox.showinfo("Export Complete", f"CSV report saved to:\n{filepath}")

        except Exception as e:
            logger.error(f"CSV export failed: {str(e)}", exc_info=True)
            messagebox.showerror("Export Failed", f"Failed to export CSV:\n{str(e)}")

    def export_report_excel(self) -> None:
        """Export current filtered report to Excel (xlsx)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except Exception:
            messagebox.showerror(
                "Missing Dependency",
                "openpyxl is required for Excel export.\n\nInstall it with:\npip install openpyxl"
            )
            return

        try:
            config.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = config.EXPORTS_DIR / f"report_{timestamp}.xlsx"

            data = self._collect_report_export_data()

            wb = Workbook()
            ws_summary = wb.active if wb.active else wb.create_sheet(title="Summary")
            ws_summary.title = "Summary"

            bold = Font(bold=True)

            ws_summary.append(["Filters"])
            if ws_summary:
                ws_summary["A1"].font = bold
            ws_summary.append(["Equities", data['filters']['equities']])
            ws_summary.append(["From Date", data['filters']['from_date'] or "-"])
            ws_summary.append(["To Date", data['filters']['to_date'] or "-"])
            ws_summary.append(["Type1", data['filters']['type1'] or "All"])
            ws_summary.append(["Expiry Month", data['filters']['expiry_month'] or "-"])
            ws_summary.append(["Include Open Positions", str(data['filters']['include_open_positions'])])
            ws_summary.append([])
            ws_summary.append(["Summary"])
            if ws_summary:
                ws_summary["A7"].font = bold
            ws_summary.append(["Total Profit", data['summary']['total_profit']])
            ws_summary.append(["Total Loss", data['summary']['total_loss']])
            ws_summary.append(["Net P/L", data['summary']['net_pnl']])
            ws_summary.append(["Win/Loss Ratio", data['summary']['win_loss_ratio']])
            ws_summary.append(["Win Rate", data['summary']['win_rate']])
            ws_summary.append(["Profit Factor", data['summary']['profit_factor']])
            ws_summary.append(["Avg Win", data['summary']['avg_win']])
            ws_summary.append(["Avg Loss", data['summary']['avg_loss']])
            ws_summary.append(["Expectancy", data['summary']['expectancy']])
            ws_summary.append(["Avg Holding Period (days)", data['summary']['avg_holding_days']])
            ws_summary.append(["Median Holding Period (days)", data['summary']['median_holding_days']])
            ws_summary.append(["Max Drawdown", data['summary']['max_drawdown']])

            ws_period = wb.create_sheet(title=f"{data['period']['type']} PnL")
            ws_period.append(["Period", "Profit", "Loss", "Net P/L", "Running Total"])
            for cell in ws_period[1]:
                cell.font = bold
            for row in data['period']['rows']:
                ws_period.append(list(row))

            ws_equity = wb.create_sheet(title="Equity Summary")
            ws_equity.append(["Equity", "Closed P/L", "Open P/L", "Total"])
            for cell in ws_equity[1]:
                cell.font = bold
            for row in data['equity_summary']:
                ws_equity.append(list(row))

            if data['open_positions']:
                ws_open = wb.create_sheet(title="Open Positions")
                ws_open.append([
                    "Equity", "Type1", "Type2", "Strike", "Expiry",
                    "Holding Days", "Status", "Qty", "Avg Price", "Unrealized P/L"
                ])
                for cell in ws_open[1]:
                    cell.font = bold
                for row in data['open_positions']:
                    ws_open.append(list(row))

            wb.save(filepath)
            messagebox.showinfo("Export Complete", f"Excel report saved to:\n{filepath}")

        except Exception as e:
            logger.error(f"Excel export failed: {str(e)}", exc_info=True)
            messagebox.showerror("Export Failed", f"Failed to export Excel:\n{str(e)}")

    def on_tab_selected(self) -> None:
        """Called when Reports tab is selected. Triggers recalculation."""
        logger.info("Reports tab selected - triggering automatic recalculation")
        self.calculate_reports()

    def apply_filters(self) -> None:
        """Apply filters and recalculate reports."""
        logger.info("Applying filters...")

        from_date = self.from_date_entry.get().strip() if hasattr(self, 'from_date_entry') else self.from_date_var.get().strip()
        to_date = self.to_date_entry.get().strip() if hasattr(self, 'to_date_entry') else self.to_date_var.get().strip()
        expiry_month = self.expiry_filter_entry.get().strip() if hasattr(self, 'expiry_filter_entry') else self.expiry_month_var.get().strip()

        if from_date and not self._validate_date_format(from_date):
            messagebox.showerror("Invalid Date", "From Date must be in YYYY-MM-DD format")
            return

        if to_date and not self._validate_date_format(to_date):
            messagebox.showerror("Invalid Date", "To Date must be in YYYY-MM-DD format")
            return

        if expiry_month and not self._validate_month_format(expiry_month):
            messagebox.showerror("Invalid Month", "Expiry month must be in YYYY-MM format")
            return

        if from_date and to_date and from_date > to_date:
            messagebox.showerror("Invalid Range", "From Date cannot be after To Date")
            return

        self.calculate_reports()

    def apply_expiry_preset(self) -> None:
        """Apply expiry month preset selection."""
        preset = self.expiry_preset_var.get().strip().lower()
        if preset == "current month":
            month_str = datetime.now().strftime("%Y-%m")
        elif preset == "next month":
            today = datetime.now()
            year = today.year + (1 if today.month == 12 else 0)
            month = 1 if today.month == 12 else today.month + 1
            month_str = f"{year}-{month:02d}"
        else:
            return

        self.expiry_month_var.set(month_str)
        if hasattr(self, 'expiry_filter_entry'):
            self.expiry_filter_entry.delete(0, 'end')
            self.expiry_filter_entry.insert(0, month_str)

    def clear_filters(self) -> None:
        """Clear all filters and recalculate."""
        logger.info("Clearing all filters")
        self.from_date_var.set("")
        self.to_date_var.set("")
        self.type1_filter_var.set("All")
        self.expiry_month_var.set("")
        self.expiry_preset_var.set("Custom")
        if hasattr(self, 'from_date_entry'):
            self.from_date_entry.delete(0, 'end')
        if hasattr(self, 'to_date_entry'):
            self.to_date_entry.delete(0, 'end')
        if hasattr(self, 'expiry_filter_entry'):
            self.expiry_filter_entry.delete(0, 'end')
        self.show_open_positions_var.set(False)

        self.select_all_equities()
        self.select_all_profiles()

        if hasattr(self, 'open_positions_frame'):
            self.open_positions_frame.pack_forget()

        self.calculate_reports()

    def select_all_equities(self) -> None:
        """Select all equities in the listbox (includes 'All')."""
        if not hasattr(self, 'equity_listbox'):
            return
        self.equity_listbox.selection_clear(0, tk.END)
        if self.equity_listbox.size() > 0:
            self.equity_listbox.selection_set(0)

    def select_all_profiles(self) -> None:
        """Select all profiles in the listbox (includes 'All')."""
        if not hasattr(self, 'profile_listbox'):
            return
        self.profile_listbox.selection_clear(0, tk.END)
        if self.profile_listbox.size() > 0:
            self.profile_listbox.selection_set(0)

    def clear_equity_selection(self) -> None:
        """Clear equity listbox selection."""
        if not hasattr(self, 'equity_listbox'):
            return
        self.equity_listbox.selection_clear(0, tk.END)

    def clear_profile_selection(self) -> None:
        """Clear profile listbox selection."""
        if not hasattr(self, 'profile_listbox'):
            return
        self.profile_listbox.selection_clear(0, tk.END)

    def _get_selected_equities(self) -> list[str] | None:
        """Get selected equities or None for all."""
        if not hasattr(self, 'equity_listbox'):
            return None
        selected = [self.equity_listbox.get(i) for i in self.equity_listbox.curselection()]
        if not selected or "All" in selected:
            return None
        return selected

    def _get_selected_profiles(self) -> list[str] | None:
        """Get selected profiles or None for all."""
        if not hasattr(self, 'profile_listbox'):
            return None
        selected = [self.profile_listbox.get(i) for i in self.profile_listbox.curselection()]
        if not selected or "All" in selected:
            return None
        return selected

    def _set_equity_listbox_values(self, equities: list[str]) -> None:
        """Populate equity listbox while preserving selection when possible."""
        if not hasattr(self, 'equity_listbox'):
            return
        current = set(self._get_selected_equities() or [])
        self.equity_listbox.delete(0, tk.END)
        for equity in equities:
            self.equity_listbox.insert(tk.END, equity)
        if not current:
            self.select_all_equities()
            return
        for idx, equity in enumerate(equities):
            if equity in current:
                self.equity_listbox.selection_set(idx)

    def _set_profile_listbox_values(self, profiles: list[str]) -> None:
        """Populate profile listbox while preserving selection when possible."""
        if not hasattr(self, 'profile_listbox'):
            return
        current = set(self._get_selected_profiles() or [])
        self.profile_listbox.delete(0, tk.END)
        for profile in profiles:
            self.profile_listbox.insert(tk.END, profile)
        if not current:
            self.select_all_profiles()
            return
        for idx, profile in enumerate(profiles):
            if profile in current:
                self.profile_listbox.selection_set(idx)

    def _refresh_profile_breakdown(self) -> None:
        """Refresh the profile breakdown tree based on selected metric."""
        if not hasattr(self, 'profile_tree'):
            return

        self.profile_tree.delete(*self.profile_tree.get_children())
        if not self.profile_breakdown_data:
            return

        mode = self.profile_breakdown_mode_var.get().strip().lower()
        if mode.startswith("profit"):
            metric_key = "profit"
            heading = "Profit (₹)"
        elif mode.startswith("loss"):
            metric_key = "loss"
            heading = "Loss (₹)"
        else:
            metric_key = "net"
            heading = "Net P/L (₹)"

        self.profile_tree.heading('Value', text=heading)

        rows = sorted(
            self.profile_breakdown_data.items(),
            key=lambda item: item[1].get(metric_key, 0),
            reverse=True
        )
        for pname, metrics in rows:
            self.profile_tree.insert('', 'end', values=(
                pname,
                format_money(metrics.get(metric_key, 0))
            ))

    def _filter_open_positions(self, positions: Sequence[OpenPosition]) -> list[OpenPosition]:
        """Filter open positions by Type1 and expiry month."""
        filtered = list(positions)
        type1_filter = self.type1_filter_var.get().strip().lower()
        if type1_filter and type1_filter != "all":
            filtered = [
                pos for pos in filtered
                if (pos.get('type1') or "delivery") == type1_filter
            ]

        expiry_month = self.expiry_filter_entry.get().strip() if hasattr(self, 'expiry_filter_entry') else self.expiry_month_var.get().strip()
        if expiry_month:
            filtered = [pos for pos in filtered if (pos.get('expiry') or "").startswith(expiry_month)]

        return filtered

    def toggle_open_positions_display(self) -> None:
        """Toggle display of open positions table."""
        if self.show_open_positions_var.get():
            self.show_open_positions_table()
        else:
            self.hide_open_positions_table()

    def show_open_positions_table(self) -> None:
        """Show open positions table."""
        if not hasattr(self, 'open_positions_frame'):
            self.open_positions_frame = ttk.LabelFrame(
                self.main_frame,
                text="📦 Open Positions",
                style="Report.TLabelframe"
            )

            table_frame = ttk.Frame(self.open_positions_frame)
            table_frame.pack(fill='both', expand=True)

            scrollbar = ttk.Scrollbar(table_frame)
            scrollbar.pack(side='right', fill='y')

            columns = (
                'Equity', 'Type1', 'Type2', 'Strike', 'Expiry',
                'Holding Days', 'Status', 'Qty', 'Avg Price', 'Unrealized P/L'
            )
            self.open_tree = ttk.Treeview(
                table_frame,
                columns=columns,
                show='headings',
                height=8,
                yscrollcommand=scrollbar.set
            )
            scrollbar.config(command=self.open_tree.yview)

            self.open_tree.heading('Equity', text='Equity', anchor='w')
            self.open_tree.heading('Type1', text='Type1', anchor='center')
            self.open_tree.heading('Type2', text='Type2', anchor='center')
            self.open_tree.heading('Strike', text='Strike', anchor='e')
            self.open_tree.heading('Expiry', text='Expiry', anchor='center')
            self.open_tree.heading('Holding Days', text='Holding Days', anchor='center')
            self.open_tree.heading('Status', text='Status', anchor='center')
            self.open_tree.heading('Qty', text='Quantity', anchor='e')
            self.open_tree.heading('Avg Price', text='Avg Price (₹)', anchor='e')
            self.open_tree.heading('Unrealized P/L', text='Unrealized P/L (₹)', anchor='e')

            self.open_tree.column('Equity', width=120, anchor='w')
            self.open_tree.column('Type1', width=90, anchor='center')
            self.open_tree.column('Type2', width=60, anchor='center')
            self.open_tree.column('Strike', width=80, anchor='e')
            self.open_tree.column('Expiry', width=100, anchor='center')
            self.open_tree.column('Holding Days', width=110, anchor='center')
            self.open_tree.column('Status', width=80, anchor='center')
            self.open_tree.column('Qty', width=100, anchor='e')
            self.open_tree.column('Avg Price', width=120, anchor='e')
            self.open_tree.column('Unrealized P/L', width=150, anchor='e')

            self.open_tree.tag_configure('evenrow', background='#f8f9fa')
            self.open_tree.tag_configure('oddrow', background='white')

            self.open_tree.pack(fill='both', expand=True)

        self.update_open_positions_table()
        self.open_positions_frame.pack(fill='x', pady=(10, 0))

    def hide_open_positions_table(self) -> None:
        """Hide open positions table."""
        if hasattr(self, 'open_positions_frame'):
            self.open_positions_frame.pack_forget()

    def update_open_positions_table(self) -> None:
        """Update open positions table with current data."""
        if not hasattr(self, 'open_tree'):
            return

        for item in self.open_tree.get_children():
            self.open_tree.delete(item)

        filtered_positions = self.filtered_open_positions
        selected_equities = self._get_selected_equities()
        if selected_equities:
            filtered_positions = [p for p in filtered_positions if p['equity'] in selected_equities]

        if not filtered_positions:
            self.open_tree.insert('', 'end', values=(
                "No open positions",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                ""
            ), tags=('empty',))
            self.open_tree.tag_configure('empty', foreground='gray')
            return

        for idx, pos in enumerate(filtered_positions):
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'

            unrealized_display = format_money(pos['unrealized_pnl']) if pos['unrealized_pnl'] != 0 else "₹ 0.00 (no market data)"

            type1_display = (pos.get('type1') or 'delivery').upper()
            type2_display = pos.get('type2') or ""
            strike_display = f"{pos['strike']:.2f}" if pos.get('strike') is not None else ""
            expiry_display = ""
            if pos.get('expiry'):
                year_e, month_e, day_e = pos['expiry'].split('-')
                expiry_display = f"{day_e}-{month_e}-{year_e}"
            holding_display = str(pos.get('holding_days', 0))

            self.open_tree.insert('', 'end', values=(
                pos['equity'],
                type1_display,
                type2_display,
                strike_display,
                expiry_display,
                holding_display,
                pos['status'],
                f"{pos['remaining_qty']:,}",
                f"₹ {pos['avg_price']:.2f}",
                unrealized_display
            ), tags=(tag,))

    def update_equity_summary(self) -> None:
        """Update equity-wise P/L summary table."""
        if not hasattr(self, 'equity_tree'):
            return

        for item in self.equity_tree.get_children():
            self.equity_tree.delete(item)

        if not self.equity_pnl and not self.filtered_open_positions:
            self.equity_tree.insert('', 'end', values=(
                "No data",
                "",
                "",
                ""
            ), tags=('empty',))
            self.equity_tree.tag_configure('empty', foreground='gray')
            return

        equity_data = {}

        for equity, pnl_data in self.equity_pnl.items():
            equity_data[equity] = {
                'closed_pnl': pnl_data['net'],
                'open_pnl': 0,
                'has_open': False
            }

        for pos in self.filtered_open_positions:
            equity = pos['equity']
            if equity not in equity_data:
                equity_data[equity] = {
                    'closed_pnl': 0,
                    'open_pnl': pos['unrealized_pnl'],
                    'has_open': True
                }
            else:
                equity_data[equity]['open_pnl'] += pos['unrealized_pnl']
                equity_data[equity]['has_open'] = True

        selected_equities = self._get_selected_equities()
        if selected_equities:
            equity_data = {k: v for k, v in equity_data.items() if k in selected_equities}

        for idx, (equity, data) in enumerate(sorted(equity_data.items())):
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'

            closed_pnl = data['closed_pnl']
            open_pnl = data['open_pnl']
            total = closed_pnl + open_pnl

            open_pnl_display = format_money(open_pnl) if open_pnl != 0 else "₹ 0.00 (no market data)"

            self.equity_tree.insert('', 'end', values=(
                equity,
                format_money(closed_pnl),
                open_pnl_display,
                format_money(total)
            ), tags=(tag,))

    def _weighted_median(self, buckets: list[tuple[int, int]]) -> float:
        """Compute weighted median from (value, weight) pairs."""
        total_weight = sum(weight for _value, weight in buckets)
        if total_weight == 0:
            return 0.0
        sorted_buckets = sorted(buckets, key=lambda x: x[0])
        running = 0
        midpoint = total_weight / 2
        for value, weight in sorted_buckets:
            running += weight
            if running >= midpoint:
                return float(value)
        return float(sorted_buckets[-1][0])

    def _calculate_max_drawdown(self, daily_pnl_totals: dict[str, int]) -> int:
        """Calculate max drawdown from daily realized P/L totals."""
        if not daily_pnl_totals:
            return 0
        cumulative = 0
        peak = 0
        max_drawdown = 0
        for date_key in sorted(daily_pnl_totals.keys()):
            cumulative += daily_pnl_totals[date_key]
            if cumulative > peak:
                peak = cumulative
            drawdown = cumulative - peak
            if drawdown < max_drawdown:
                max_drawdown = drawdown
        return max_drawdown

    def _validate_date_format(self, date_str: str) -> bool:
        """Validate date string is in YYYY-MM-DD format."""
        try:
            datetime.strptime(date_str, '%Y-%m-%d')
            return True
        except ValueError:
            return False

    def _validate_month_format(self, month_str: str) -> bool:
        """Validate month string is in YYYY-MM format."""
        try:
            datetime.strptime(month_str, '%Y-%m')
            return True
        except ValueError:
            return False
