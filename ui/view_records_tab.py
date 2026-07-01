"""
View Records Tab - Phase 2

Responsibilities:
- Fetch trades from database
- Display all trades in table
- Filter by equity, trade type, date range
- Allow edit trade (update existing record)
- Allow soft delete (set is_active = 0)

Does NOT:
- Calculate FIFO
- Calculate P/L
- Prevent oversells
- Auto-fix data
"""

import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import sqlite3
import csv
from pathlib import Path
from typing import Callable, Optional
from core.logger import get_logger
from core.utils import format_money
from core.trade_validation import normalize_trade_classification
from core.brokerage import calculate_brokerage_auto
from core.utils import make_trade_ts
import config

logger = get_logger('ui.view_records_tab')

CALENDAR_AVAILABLE: bool
try:
    from tkcalendar import DateEntry  # type: ignore
    CALENDAR_AVAILABLE = True
except ImportError:
    DateEntry = None  # type: ignore
    CALENDAR_AVAILABLE = False
    logger.warning("tkcalendar not installed - using text entry for dates")


class ViewRecordsTab:
    """View Records tab - display and manage trade records."""
    
    def __init__(self, parent: ttk.Frame, status_callback: Callable[[str], None]) -> None:
        logger.info("Initializing View Records tab")
        self.parent = parent
        self.update_status = status_callback
        self.selected_trade_id: Optional[int] = None
        self.show_deleted_trades = tk.BooleanVar(value=False)
        
        # Create UI
        self.create_widgets()
        
        # Load initial data
        self.refresh_records()
        logger.debug("View Records tab initialized")
    
    def create_widgets(self) -> None:
        """Create all UI widgets for View Records tab."""
        
        # Main container
        main_frame = ttk.Frame(self.parent, padding="10")
        main_frame.pack(fill='both', expand=True)
        
        # Header
        header = ttk.Label(
            main_frame,
            text="VIEW TRADE RECORDS",
            font=('Consolas', 16, 'bold')
        )
        header.pack(pady=(0, 10))
        
        # Filter section
        self.create_filter_section(main_frame)
        
        # Table section
        self.create_table_section(main_frame)
        
        # Button section
        self.create_button_section(main_frame)
    
    def create_filter_section(self, parent: ttk.Frame) -> None:
        """Create filter controls."""
        
        filter_frame = ttk.LabelFrame(parent, text="Filters", padding="10")
        filter_frame.pack(fill='x', pady=(0, 10))
        
        # Row 1: Equity and Trade Type
        row1 = ttk.Frame(filter_frame)
        row1.pack(fill='x', pady=5)
        
        # Equity filter
        ttk.Label(row1, text="Stock:", font=('Arial', 9)).pack(side='left', padx=(0, 5))
        self.equity_filter = ttk.Combobox(row1, width=15, state='readonly')
        self.equity_filter.pack(side='left', padx=(0, 20))
        self.equity_filter.set("All")
        
        # Trade Type filter
        ttk.Label(row1, text="Type:", font=('Arial', 9)).pack(side='left', padx=(0, 5))
        self.type_filter = ttk.Combobox(row1, width=10, state='readonly', values=["All", "BUY", "SELL"])
        self.type_filter.pack(side='left', padx=(0, 20))
        self.type_filter.set("All")
        
        # Show deleted toggle
        self.show_deleted_check = ttk.Checkbutton(
            row1,
            text="Show Deleted",
            variable=self.show_deleted_trades,
            command=self.refresh_records
        )
        self.show_deleted_check.pack(side='left', padx=(0, 20))
        
        # Apply button
        ttk.Button(row1, text="Apply Filters", command=self.apply_filters, width=12).pack(side='left', padx=10)
        ttk.Button(row1, text="Clear Filters", command=self.clear_filters, width=12).pack(side='left')
        
        # Row 2: Date Range Filter
        row2 = ttk.Frame(filter_frame)
        row2.pack(fill='x', pady=5)
        
        ttk.Label(row2, text="From:", font=('Arial', 9)).pack(side='left', padx=(0, 5))
        
        if CALENDAR_AVAILABLE:
            # Use calendar picker (start empty for filtering)
            date_frame_from = ttk.Frame(row2)
            date_frame_from.pack(side='left', padx=(0, 15))
            
            self.date_from_entry = DateEntry(
                date_frame_from,
                width=12,
                background='darkblue',
                foreground='white',
                borderwidth=2,
                date_pattern='dd-mm-yyyy',
                font=('Arial', 9)
            )
            self.date_from_entry.pack(side='left')
            # Clear the default date (make it empty)
            self.date_from_entry.delete(0, 'end')
        else:
            # Fallback to text entry (start empty)
            self.date_from_entry = ttk.Entry(row2, width=12)
            self.date_from_entry.pack(side='left', padx=(0, 5))
            self.date_from_entry.bind('<FocusOut>', lambda e: self.validate_date_field(self.date_from_entry))
            ttk.Label(row2, text="(DD-MM-YYYY)", font=('Arial', 8), foreground='gray').pack(side='left', padx=(0, 15))
        
        ttk.Label(row2, text="To:", font=('Arial', 9)).pack(side='left', padx=(0, 5))
        
        if CALENDAR_AVAILABLE:
            # Use calendar picker (start empty for filtering)
            date_frame_to = ttk.Frame(row2)
            date_frame_to.pack(side='left')
            
            self.date_to_entry = DateEntry(
                date_frame_to,
                width=12,
                background='darkblue',
                foreground='white',
                borderwidth=2,
                date_pattern='dd-mm-yyyy',
                font=('Arial', 9)
            )
            self.date_to_entry.pack(side='left')
            # Clear the default date (make it empty)
            self.date_to_entry.delete(0, 'end')
        else:
            # Fallback to text entry
            self.date_to_entry = ttk.Entry(row2, width=12)
            self.date_to_entry.pack(side='left', padx=(0, 5))
            self.date_to_entry.bind('<FocusOut>', lambda e: self.validate_date_field(self.date_to_entry))
            ttk.Label(row2, text="(DD-MM-YYYY)", font=('Arial', 8), foreground='gray').pack(side='left')
    
    def create_table_section(self, parent: ttk.Frame) -> None:
        """Create table to display records."""
        
        # Table frame with scrollbars
        table_frame = ttk.Frame(parent)
        table_frame.pack(fill='both', expand=True, pady=(0, 10))
        
        # Scrollbars
        vsb = ttk.Scrollbar(table_frame, orient="vertical")
        hsb = ttk.Scrollbar(table_frame, orient="horizontal")
        
        # Treeview
        columns = (
            'ID', 'Date', 'Stock', 'Type', 'Type1', 'Type2', 'Strike', 'Expiry',
            'Qty', 'Price', 'Brokerage', 'Mtf_Amt', 'Mtf_Rate', 'Notes', 'Status'
        )
        self.records_tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show='headings',
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set
        )
        
        vsb.config(command=self.records_tree.yview)
        hsb.config(command=self.records_tree.xview)
        
        # Configure columns with clickable headings for sorting
        self._records_sort_state = {}
        def _make_heading(col_name, display_text=None):
            text = display_text or col_name
            self.records_tree.heading(col_name, text=text, command=lambda c=col_name: self._on_records_heading_click(c))
            self._records_sort_state[col_name] = None

        _make_heading('ID', 'ID')
        _make_heading('Date', 'Date')
        _make_heading('Stock', 'Stock')
        _make_heading('Type', 'Type')
        _make_heading('Type1', 'Type1')
        _make_heading('Type2', 'Type2')
        _make_heading('Strike', 'Strike')
        _make_heading('Expiry', 'Expiry')
        _make_heading('Qty', 'Qty')
        _make_heading('Price', 'Price (₹)')
        _make_heading('Brokerage', 'Brokerage (₹)')
        _make_heading('Mtf_Amt', 'MTF Amt (₹)')
        _make_heading('Mtf_Rate', 'MTF Rate (%)')
        _make_heading('Notes', 'Notes')
        _make_heading('Status', 'Status')
        
        self.records_tree.column('ID', width=50, anchor='center')
        self.records_tree.column('Date', width=100, anchor='center')
        self.records_tree.column('Stock', width=80, anchor='center')
        self.records_tree.column('Type', width=60, anchor='center')
        self.records_tree.column('Type1', width=90, anchor='center')
        self.records_tree.column('Type2', width=60, anchor='center')
        self.records_tree.column('Strike', width=80, anchor='e')
        self.records_tree.column('Expiry', width=100, anchor='center')
        self.records_tree.column('Qty', width=60, anchor='center')
        self.records_tree.column('Price', width=100, anchor='e')
        self.records_tree.column('Brokerage', width=100, anchor='e')
        self.records_tree.column('Mtf_Amt', width=90, anchor='e')
        self.records_tree.column('Mtf_Rate', width=90, anchor='e')
        self.records_tree.column('Notes', width=200, anchor='w')
        self.records_tree.column('Status', width=80, anchor='center')
        
        # Grid layout
        self.records_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        # Bind double-click to inline edit
        self.records_tree.bind('<Double-Button-1>', self.on_double_click)
        
        # Bind selection
        self.records_tree.bind('<<TreeviewSelect>>', self.on_select)
        
        # Bind mousewheel for scrolling
        self.records_tree.bind('<MouseWheel>', lambda e: self.records_tree.yview_scroll(int(-1*(e.delta/120)), "units"))
    
    def create_button_section(self, parent: ttk.Frame) -> None:
        """Create action buttons."""
        
        button_frame = ttk.Frame(parent)
        button_frame.pack(fill='x')
        
        ttk.Button(
            button_frame,
            text="Refresh",
            command=self.refresh_records,
            width=15
        ).pack(side='left', padx=5)
        
        ttk.Button(
            button_frame,
            text="✏️ Edit Trade",
            command=self.edit_selected_trade,
            width=15
        ).pack(side='left', padx=5)
        
        ttk.Button(
            button_frame,
            text="Delete Trade",
            command=self.delete_selected_trade,
            width=15
        ).pack(side='left', padx=5)

        ttk.Button(
            button_frame,
            text="Restore Trade",
            command=self.restore_selected_trade,
            width=15
        ).pack(side='left', padx=5)
        
        # Separator
        ttk.Separator(button_frame, orient='vertical').pack(side='left', fill='y', padx=10)
        
        ttk.Button(
            button_frame,
            text="💾 Backup DB",
            command=self.backup_database,
            width=15
        ).pack(side='left', padx=5)
        
        ttk.Button(
            button_frame,
            text="📥 Restore DB",
            command=self.restore_database,
            width=15
        ).pack(side='left', padx=5)
        
        # Separator
        ttk.Separator(button_frame, orient='vertical').pack(side='left', fill='y', padx=10)
        
        ttk.Button(
            button_frame,
        

            text="📥 Import CSV",
            command=self.import_trades_csv,
            width=15
        ).pack(side='left', padx=5)
        
        ttk.Button(
            button_frame,
            text="ℹ️ CSV Format",
            command=self.show_import_format,
            width=15
        ).pack(side='left', padx=5)
        
        ttk.Button(
            button_frame,
            text="💾 Export CSV",
            
            command=self.export_to_csv,
            width=15
        ).pack(side='left', padx=5)
        
        ttk.Button(
            button_frame,
            text="📄 Export Excel",
            command=self.export_to_excel,
            width=15
        ).pack(side='left', padx=5)
        
        # Info label
        self.info_label = ttk.Label(button_frame, text="", font=('Arial', 9), foreground='blue')
        self.info_label.pack(side='right', padx=10)
    
    def load_equity_list(self) -> None:
        """Load unique equity list for filter dropdown."""
        try:
            conn = sqlite3.connect(str(config.DB_PATH))
            c = conn.cursor()
            # Apply profile filter when loading equities
            active_ids = config.ACTIVE_PROFILE_IDS
            if not active_ids:
                c.execute("SELECT DISTINCT equity FROM trade_events ORDER BY equity")
            else:
                placeholders = ','.join('?' * len(active_ids))
                c.execute(f"SELECT DISTINCT equity FROM trade_events WHERE profile_id IN ({placeholders}) ORDER BY equity", tuple(active_ids))
            equities = [row[0] for row in c.fetchall()]
            conn.close()
            
            self.equity_filter['values'] = ["All"] + equities
            logger.debug(f"Loaded {len(equities)} unique equities for filter")
        except Exception as e:
            logger.error(f"Failed to load equity list: {str(e)}", exc_info=True)
    
    def refresh_records(self) -> None:
        """Load all trade records from database."""
        logger.info("Refreshing trade records")
        
        # Clear existing items
        for item in self.records_tree.get_children():
            self.records_tree.delete(item)
        
        try:
            # Load equity list for filter
            self.load_equity_list()
            
            # Build query based on filters
            query = """
                SELECT id, trade_date, equity, trade_type, type1, type2, strike, expiry,
                       quantity, price, brokerage, mtf_amount, mtf_rate_ppm, notes, is_active
                FROM trade_events
                WHERE 1=1
            """
            params = []
            
            # Apply filters
            equity_filter = self.equity_filter.get()
            if equity_filter and equity_filter != "All":
                query += " AND equity = ?"
                params.append(equity_filter)
            
            type_filter = self.type_filter.get()
            if type_filter and type_filter != "All":
                query += " AND trade_type = ?"
                params.append(type_filter)
            
            # Date range filter
            date_from = self.date_from_entry.get().strip()
            date_to = self.date_to_entry.get().strip()
            
            if date_from:
                try:
                    # Convert DD-MM-YYYY to YYYY-MM-DD
                    day, month, year = date_from.split('-')
                    date_from_db = f"{year}-{month}-{day}"
                    query += " AND trade_date >= ?"
                    params.append(date_from_db)
                except ValueError:
                    logger.warning(f"Invalid date format for 'from': {date_from}")
            
            if date_to:
                try:
                    # Convert DD-MM-YYYY to YYYY-MM-DD
                    day, month, year = date_to.split('-')
                    date_to_db = f"{year}-{month}-{day}"
                    query += " AND trade_date <= ?"
                    params.append(date_to_db)
                except ValueError:
                    logger.warning(f"Invalid date format for 'to': {date_to}")
            
            # Apply deleted filter based on checkbox
            if not self.show_deleted_trades.get():
                query += " AND is_active = 1"  # Only show active trades
            # If checkbox is checked, show all (active + deleted)
            # Apply profile filter if set (empty list => combined view)
            import config as _config
            active_ids = _config.ACTIVE_PROFILE_IDS
            if active_ids:
                placeholders = ','.join('?' * len(active_ids))
                query += f" AND profile_id IN ({placeholders})"
                params.extend(active_ids)

            query += " ORDER BY id DESC"
            
            # Execute query
            conn = sqlite3.connect(str(config.DB_PATH))
            c = conn.cursor()
            c.execute(query, params)
            trades = c.fetchall()
            conn.close()
            
            logger.debug(f"Loaded {len(trades)} trade records")
            
            # Populate table
            for trade in trades:
                (trade_id, trade_date, equity, trade_type, type1, type2, strike, expiry,
                 quantity, price_paise, brokerage_paise, mtf_amount_paise, mtf_rate_ppm, notes, is_active) = trade
                
                # Format date DD-MM-YYYY
                year, month, day = trade_date.split('-')
                display_date = f"{day}-{month}-{year}"
                
                # Status
                status = "Active" if is_active == 1 else "Deleted"
                
                # Normalize display values
                type1_display = (type1 or "delivery").upper()
                type2_display = type2 if type2 else ""
                strike_display = f"{strike:.2f}" if strike is not None else ""
                expiry_display = ""
                if expiry:
                    year_e, month_e, day_e = expiry.split('-')
                    expiry_display = f"{day_e}-{month_e}-{year_e}"

                mtf_rate_display = f"{(mtf_rate_ppm / 10000):.2f}" if mtf_rate_ppm is not None else ""

                # Insert into tree
                item = self.records_tree.insert('', 'end', values=(
                    trade_id,
                    display_date,
                    equity,
                    trade_type,
                    type1_display,
                    type2_display,
                    strike_display,
                    expiry_display,
                    quantity,
                    format_money(price_paise),
                    format_money(brokerage_paise),
                    format_money(mtf_amount_paise or 0),
                    mtf_rate_display,
                    notes[:50] + "..." if len(notes) > 50 else notes,
                    status
                ))
                
                # Apply visual styling
                if is_active == 0:
                    # Deleted rows: gray text
                    self.records_tree.item(item, tags=('deleted',))
                elif trade_type == 'SELL':
                    # SELL rows: light red background
                    self.records_tree.item(item, tags=('sell',))
                elif trade_type == 'BUY':
                    # BUY rows: light green background
                    self.records_tree.item(item, tags=('buy',))
            
            # Configure tag colors
            self.records_tree.tag_configure('deleted', foreground='gray')
            self.records_tree.tag_configure('sell', background='#ffe6e6')  # Light red
            self.records_tree.tag_configure('buy', background='#e6ffe6')   # Light green

            # Reset headings arrows (if any)
            for col in self._records_sort_state:
                # remove any arrow from heading text
                heading_text = col
                if col == 'Price':
                    heading_text = 'Price (₹)'
                elif col == 'Brokerage':
                    heading_text = 'Brokerage (₹)'
                self.records_tree.heading(col, text=heading_text)
            
            # Update info label
            self.info_label.config(text=f"Total: {len(trades)} records")
            self.update_status(f"Loaded {len(trades)} trade records")
            
        except Exception as e:
            logger.error(f"Failed to load records: {str(e)}", exc_info=True)
            messagebox.showerror("Error", f"Failed to load records:\n{str(e)}")
            self.update_status("❌ Error loading records")

    def _build_trade_filters(self) -> tuple[str, list]:
        """Build SQL WHERE filters from current UI filter state."""
        where_clause = " WHERE 1=1"
        params: list = []

        equity_filter = self.equity_filter.get()
        if equity_filter and equity_filter != "All":
            where_clause += " AND equity = ?"
            params.append(equity_filter)

        type_filter = self.type_filter.get()
        if type_filter and type_filter != "All":
            where_clause += " AND trade_type = ?"
            params.append(type_filter)

        date_from = self.date_from_entry.get().strip()
        if date_from:
            try:
                day, month, year = date_from.split('-')
                where_clause += " AND trade_date >= ?"
                params.append(f"{year}-{month}-{day}")
            except ValueError:
                logger.warning(f"Invalid date format for 'from': {date_from}")

        date_to = self.date_to_entry.get().strip()
        if date_to:
            try:
                day, month, year = date_to.split('-')
                where_clause += " AND trade_date <= ?"
                params.append(f"{year}-{month}-{day}")
            except ValueError:
                logger.warning(f"Invalid date format for 'to': {date_to}")

        if not self.show_deleted_trades.get():
            where_clause += " AND is_active = 1"
        # Apply profile filter if set (0 => combined view)
        import config as _config
        active_ids = _config.ACTIVE_PROFILE_IDS
        if active_ids:
            placeholders = ','.join('?' * len(active_ids))
            where_clause += f" AND profile_id IN ({placeholders})"
            params.extend(active_ids)

        return where_clause, params

    def _fetch_filtered_trades_for_export(self) -> list[tuple]:
        """Fetch filtered trade rows with full v1.1 fields for export."""
        where_clause, params = self._build_trade_filters()
        query = (
            "SELECT id, trade_date, trade_ts, equity, trade_type, type1, type2, strike, expiry, "
            "quantity, price, brokerage, brokerage_auto, brokerage_override, mtf_amount, mtf_rate_ppm, notes, is_active "
            f"FROM trade_events{where_clause} ORDER BY id DESC"
        )

        conn = sqlite3.connect(str(config.DB_PATH))
        c = conn.cursor()
        c.execute(query, params)
        rows = c.fetchall()
        conn.close()
        return rows
    
    def apply_filters(self) -> None:
        """Apply current filters and refresh."""
        date_from = self.date_from_entry.get().strip()
        date_to = self.date_to_entry.get().strip()
        logger.info(f"Applying filters - Equity: {self.equity_filter.get()}, Type: {self.type_filter.get()}, Date: {date_from} to {date_to}, Show Deleted: {self.show_deleted_trades.get()}")
        self.refresh_records()
    
    def validate_date_field(self, entry: ttk.Entry) -> bool:
        """Validate date format in entry field. Clear if invalid."""
        value = entry.get().strip()
        if not value:
            return True
        
        try:
            day, month, year = value.split('-')
            if len(day) != 2 or len(month) != 2 or len(year) != 4:
                raise ValueError("Invalid format")
            # Basic range check
            if not (1 <= int(day) <= 31 and 1 <= int(month) <= 12 and 1900 <= int(year) <= 2100):
                raise ValueError("Invalid date range")
            return True
        except (ValueError, AttributeError):
            messagebox.showwarning("Invalid Date", f"Please enter date in DD-MM-YYYY format\n(e.g., 20-01-2026)")
            entry.delete(0, tk.END)
            return False
    
    def clear_filters(self) -> None:
        """Clear all filters and refresh."""
        logger.info("Clearing all filters")
        self.equity_filter.set("All")
        self.type_filter.set("All")
        
        # Clear date fields (handle both DateEntry and Entry widgets)
        if CALENDAR_AVAILABLE:
            self.date_from_entry.set_date('')  # Clear DateEntry
            self.date_to_entry.set_date('')
        else:
            self.date_from_entry.delete(0, tk.END)
            self.date_to_entry.delete(0, tk.END)
        
        self.show_deleted_trades.set(False)
        self.refresh_records()
    
    def on_select(self, event: tk.Event) -> None:  # type: ignore
        """Handle row selection."""
        selection = self.records_tree.selection()
        if selection:
            item = self.records_tree.item(selection[0])
            self.selected_trade_id = item['values'][0]
            logger.debug(f"Selected trade ID: {self.selected_trade_id}")
    
    def edit_selected_trade(self) -> None:
        """Open edit dialog for selected trade."""
        selection = self.records_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a trade to edit")
            return
        
        item = self.records_tree.item(selection[0])
        trade_id = item['values'][0]
        status = item['values'][14]  # Status column
        
        # Prevent editing deleted trades
        if status == "Deleted":
            logger.warning(f"Attempted to edit deleted trade ID: {trade_id}")
            messagebox.showwarning(
                "Cannot Edit",
                "Cannot edit a deleted trade.\n\nPlease restore it first (feature coming soon)."
            )
            return
        
        logger.info(f"Opening edit dialog for trade ID: {trade_id}")
        
        # Fetch full trade details
        try:
            conn = sqlite3.connect(str(config.DB_PATH))
            c = conn.cursor()
            c.execute("""
                SELECT trade_date, trade_ts, equity, trade_type, type1, type2, strike, expiry,
                       quantity, price, brokerage, brokerage_auto, brokerage_override,
                       mtf_amount, mtf_rate_ppm, notes, is_active
                FROM trade_events
                WHERE id = ?
            """, (trade_id,))
            trade = c.fetchone()
            conn.close()
            
            if not trade:
                messagebox.showerror("Error", "Trade not found")
                return
            
            # Confirm editing (UX warning)
            trade_date = trade[0]
            result = messagebox.askyesno(
                "Confirm Edit",
                f"You are about to edit trade #{trade_id} from {trade_date}.\n\n"
                f"⚠️ Warning: Editing historical trades may affect your P/L calculations.\n\n"
                f"Do you want to continue?"
            )
            
            if not result:
                logger.info(f"Edit cancelled by user for trade ID: {trade_id}")
                return
            
            # Open edit dialog
            EditTradeDialog(self.parent, int(trade_id), trade, self.refresh_records, self.update_status)
            
        except Exception as e:
            logger.error(f"Failed to load trade for editing: {str(e)}", exc_info=True)
            messagebox.showerror("Error", f"Failed to load trade:\n{str(e)}")
    
    def delete_selected_trade(self) -> None:
        """Soft delete selected trade (set is_active = 0)."""
        selection = self.records_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a trade to delete")
            return
        
        item = self.records_tree.item(selection[0])
        trade_id = item['values'][0]
        trade_info = f"{item['values'][3]} {item['values'][8]} {item['values'][2]} on {item['values'][1]}"
        
        # Confirm deletion
        result = messagebox.askyesno(
            "Confirm Delete",
            f"Are you sure you want to delete this trade?\n\n{trade_info}\n\nThis will set is_active = 0 (soft delete)."
        )
        
        if not result:
            logger.info(f"Delete cancelled by user for trade ID: {trade_id}")
            return
        
        try:
            logger.info(f"Soft deleting trade ID: {trade_id}")
            conn = sqlite3.connect(str(config.DB_PATH))
            c = conn.cursor()
            c.execute("UPDATE trade_events SET is_active = 0 WHERE id = ?", (trade_id,))
            conn.commit()
            conn.close()
            
            logger.info(f"✅ Trade ID {trade_id} soft deleted successfully")
            self.update_status(f"✅ Trade deleted: {trade_info}")
            messagebox.showinfo("Success", "Trade deleted successfully")
            
            # Refresh
            self.refresh_records()
            
        except Exception as e:
            logger.error(f"Failed to delete trade: {str(e)}", exc_info=True)
            messagebox.showerror("Error", f"Failed to delete trade:\n{str(e)}")
            self.update_status("❌ Error deleting trade")

    def restore_selected_trade(self) -> None:
        """Restore a soft-deleted trade (set is_active = 1)."""
        selection = self.records_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a deleted trade to restore")
            return

        item = self.records_tree.item(selection[0])
        values = item.get('values', ())
        if len(values) < 15:
            messagebox.showwarning("Invalid Selection", "Please select a trade row to restore")
            return

        trade_id = values[0]
        status = values[14]
        trade_info = f"{values[3]} {values[8]} {values[2]} on {values[1]}"

        if status != "Deleted":
            messagebox.showinfo("Already Active", "This trade is already active.")
            return

        result = messagebox.askyesno(
            "Confirm Restore",
            f"Are you sure you want to restore this trade?\n\n{trade_info}\n\nThis will set is_active = 1."
        )

        if not result:
            logger.info(f"Restore cancelled by user for trade ID: {trade_id}")
            return

        try:
            logger.info(f"Restoring trade ID: {trade_id}")
            conn = sqlite3.connect(str(config.DB_PATH))
            c = conn.cursor()
            c.execute("UPDATE trade_events SET is_active = 1 WHERE id = ?", (trade_id,))
            conn.commit()
            conn.close()

            logger.info(f"✅ Trade ID {trade_id} restored successfully")
            self.update_status(f"✅ Trade restored: {trade_info}")
            messagebox.showinfo("Success", "Trade restored successfully")

            self.refresh_records()

        except Exception as e:
            logger.error(f"Failed to restore trade: {str(e)}", exc_info=True)
            messagebox.showerror("Error", f"Failed to restore trade:\n{str(e)}")
            self.update_status("❌ Error restoring trade")
    
    def on_double_click(self, event: tk.Event) -> None:  # type: ignore
        """Handle double-click for inline editing."""
        # Identify column and item
        region = self.records_tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        
        column = self.records_tree.identify_column(event.x)
        item = self.records_tree.identify_row(event.y)
        
        if not item:
            return
        
        # Get column name
        col_index = int(column[1:]) - 1  # #1 -> 0, #2 -> 1, etc.
        columns = (
            'ID', 'Date', 'Stock', 'Type', 'Type1', 'Type2', 'Strike', 'Expiry',
            'Qty', 'Price', 'Brokerage', 'Notes', 'Status'
        )
        col_name = columns[col_index]
        
        # Don't allow editing ID or Status columns
        if col_name in ('ID', 'Status', 'Type1', 'Type2', 'Strike', 'Expiry'):
            logger.debug(f"Inline edit: Column {col_name} is readonly")
            return
        
        # Check if trade is deleted
        values = self.records_tree.item(item, 'values')
        if values[12] == "Deleted":  # Status column
            messagebox.showwarning("Cannot Edit", "Cannot edit a deleted trade.")
            return
        
        logger.info(f"Inline edit: Starting edit for item {item}, column {col_name}")
        self.start_inline_edit(item, col_index, col_name)
    
    def start_inline_edit(self, item: str, col_index: int, col_name: str) -> None:
        """Start inline editing of a cell."""
        # Get cell position
        bbox = self.records_tree.bbox(item, col_index)
        if not bbox:
            return
        
        x, y, width, height = bbox
        current_value = self.records_tree.item(item, 'values')[col_index]
        
        # Remove currency symbols for editing
        if col_name in ('Price', 'Brokerage'):
            current_value = current_value.replace('₹', '').strip()
        
        # Create entry widget for inline editing
        if col_name == 'Notes':
            # For Notes, use a larger entry
            edit_entry = ttk.Entry(self.records_tree, width=40)
        elif col_name == 'Type':
            # For Type, use combobox
            edit_entry = ttk.Combobox(self.records_tree, width=10, values=['BUY', 'SELL'], state='readonly')
        else:
            edit_entry = ttk.Entry(self.records_tree)
        
        edit_entry.insert(0, current_value)
        edit_entry.select_range(0, tk.END)
        edit_entry.place(x=x, y=y, width=width, height=height)
        edit_entry.focus()
        
        # Save on Enter or focus loss
        def save_edit(event=None):
            new_value = edit_entry.get()
            edit_entry.destroy()
            self.update_inline_edit(item, col_index, col_name, new_value)
        
        def cancel_edit(event=None):
            edit_entry.destroy()
        
        edit_entry.bind('<Return>', save_edit)
        edit_entry.bind('<FocusOut>', save_edit)
        edit_entry.bind('<Escape>', cancel_edit)

    def _on_records_heading_click(self, col: str) -> None:
        """Toggle sort order for the given column and sort the treeview."""
        # Determine new sort order
        current = self._records_sort_state.get(col)
        reverse = False if current is None or current is False else True
        # Toggle
        self._records_sort_state[col] = not reverse

        # Fetch items and sort
        items = list(self.records_tree.get_children(''))

        def _val(item):
            v = self.records_tree.set(item, col)
            if col in ('ID', 'Qty'):
                try:
                    return int(v)
                except Exception:
                    return 0
            if col in ('Price', 'Brokerage'):
                try:
                    s = v.replace('₹', '').replace(',', '').strip()
                    return float(s)
                except Exception:
                    return 0.0
            if col == 'Date':
                try:
                    # Display date is DD-MM-YYYY
                    day, month, year = v.split('-')
                    return datetime.strptime(f"{year}-{month}-{day}", '%Y-%m-%d')
                except Exception:
                    return datetime.min
            return v.lower() if isinstance(v, str) else v

        items.sort(key=_val, reverse=self._records_sort_state[col])

        # Rearrange
        for index, iid in enumerate(items):
            self.records_tree.move(iid, '', index)

        # Update heading visuals (arrow)
        arrow = ' ▲' if self._records_sort_state[col] else ' ▼'
        # Reset all headings
        for c in self.records_tree['columns']:
            text = c
            if c == 'Price':
                text = 'Price (₹)'
            elif c == 'Brokerage':
                text = 'Brokerage (₹)'
            elif c == 'Mtf_Amt':
                text = 'MTF Amt (₹)'
            self.records_tree.heading(c, text=text)
        # Set arrow
        display = col
        if col == 'Price':
            display = 'Price (₹)'
        elif col == 'Brokerage':
            display = 'Brokerage (₹)'
        elif col == 'Mtf_Amt':
            display = 'MTF Amt (₹)'
        self.records_tree.heading(col, text=display + arrow)
    
    def update_inline_edit(self, item: str, col_index: int, col_name: str, new_value: str) -> None:
        """Update database with inline edited value."""
        try:
            values = self.records_tree.item(item, 'values')
            trade_id = values[0]
            
            logger.info(f"Inline edit: Updating trade ID {trade_id}, column {col_name} to '{new_value}'")
            
            # Map column name to database field
            column_map = {
                'Date': ('trade_date', 'date'),
                'Stock': ('equity', 'text'),
                'Type': ('trade_type', 'text'),
                'Qty': ('quantity', 'int'),
                'Price': ('price', 'money'),
                'Brokerage': ('brokerage', 'money'),
                'Mtf_Amt': ('mtf_amount', 'money'),
                'Mtf_Rate': ('mtf_rate_ppm', 'rate'),
                'Notes': ('notes', 'text')
            }
            
            if col_name not in column_map:
                return
            
            db_field, value_type = column_map[col_name]
            
            # Convert value based on type
            if value_type == 'date':
                # DD-MM-YYYY -> YYYY-MM-DD
                day, month, year = new_value.split('-')
                db_value = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
            elif value_type == 'int':
                db_value = int(new_value)
            elif value_type == 'money':
                # Convert rupees to paise
                db_value = int(float(new_value) * 100)
            elif value_type == 'rate':
                # Convert % to ppm
                db_value = int(float(new_value) * 10000)
            elif value_type == 'text':
                db_value = new_value.strip().upper() if col_name in ('Stock', 'Type') else new_value.strip()
            else:
                db_value = new_value
            
            # Update database
            conn = sqlite3.connect(str(config.DB_PATH))
            c = conn.cursor()

            if col_name == 'Brokerage':
                # Inline brokerage edit is treated as explicit override
                c.execute(
                    "UPDATE trade_events SET brokerage = ?, brokerage_auto = 0, brokerage_override = ? WHERE id = ?",
                    (db_value, db_value, trade_id)
                )
            else:
                c.execute(f"UPDATE trade_events SET {db_field} = ? WHERE id = ?", (db_value, trade_id))

            conn.commit()
            conn.close()
            
            logger.info(f"✅ Inline edit: Trade ID {trade_id} updated successfully")
            self.update_status(f"✅ Updated trade #{trade_id}")
            
            # Refresh display
            self.refresh_records()
            
        except Exception as e:
            logger.error(f"Inline edit failed: {str(e)}", exc_info=True)
            messagebox.showerror("Error", f"Failed to update:\n{str(e)}")
    
    def show_import_format(self) -> None:
        """Show CSV import format help."""
        format_msg = f"""CSV IMPORT FORMAT

Required Columns:
• Date (DD-MM-YYYY or YYYY-MM-DD)
• Stock (symbol, e.g., TCS, RELIANCE)
• Type (BUY or SELL)
• Qty (positive integer)
• Price (in rupees, e.g., 350.50)

Optional Columns:
• Brokerage (in rupees, default 0)
• BrokerageOverride (in rupees, optional)
• MtfAmount (in rupees, required for MTF BUY only)
• MtfRate (in %, e.g. 9.65, optional for MTF BUY)
• TradeTS (YYYY-MM-DD HH:MM:SS, optional IST timestamp)
• Notes (any text)
• Type1 (intraday/delivery/mtf/futures/options)
• Type2 (CE/PE, required for options only)
• Strike (required for options only)
• Expiry (required for options/futures)

Example CSV:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Date,Stock,Type,Qty,Price,Brokerage,BrokerageOverride,MtfAmount,MtfRate,TradeTS,Notes,Type1,Type2,Strike,Expiry
20-01-2026,TCS,BUY,10,350.50,10.00,,0,,2026-01-20 09:30:00,Sample trade,delivery,,,
21-01-2026,RELIANCE,BUY,5,280.00,,8.00,0,,2026-01-21 10:05:00,Another trade,intraday,,,
22-01-2026,NIFTY,BUY,50,12.00,2.00,,0,2026-01-22 11:00:00,Options entry,options,CE,22500,25-01-2026
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Sample file available at:
{str(config.SAMPLE_CSV_PATH)}

Tips:
✓ First row must be header
✓ No empty rows
✓ Use commas to separate values
✓ Dates validated automatically
✓ Errors shown after import"""
        
        messagebox.showinfo("CSV Import Format", format_msg)
        logger.info("Displayed CSV import format help")
    
    def import_trades_csv(self) -> None:
        """Import trades from CSV file."""
        from tkinter import filedialog
        
        logger.info("Starting CSV import")
        
        # Ask for CSV file
        file_path = filedialog.askopenfilename(
            title="Select CSV file to import",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            defaultextension=".csv"
        )
        
        if not file_path:
            logger.debug("CSV import cancelled by user")
            return
        
        try:
            imported_count = 0
            skipped_count = 0
            error_lines = []
            
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                
                # Expected columns: Date, Stock, Type, Qty, Price (Type1+ derivatives are optional)
                required_cols = ['Date', 'Stock', 'Type', 'Qty', 'Price']
                
                if not reader.fieldnames or not all(col in reader.fieldnames for col in required_cols):
                    messagebox.showerror(
                        "Invalid CSV",
                        f"CSV must contain columns: {', '.join(required_cols)}\n\n"
                        f"Found columns: {', '.join(reader.fieldnames or [])}"
                    )
                    logger.warning(f"Invalid CSV format: missing required columns")
                    return
                
                try:
                    profile_id = int(config.PRIMARY_PROFILE_ID) if config.PRIMARY_PROFILE_ID is not None else None
                except Exception:
                    profile_id = None
                if profile_id is None:
                    messagebox.showwarning(
                        "Select Profile",
                        "Multiple profiles or Combined Family view selected. Select a single profile first."
                    )
                    return

                with sqlite3.connect(str(config.DB_PATH), timeout=10) as conn:
                    conn.execute("PRAGMA foreign_keys = ON")
                    cursor = conn.cursor()

                    for line_num, row in enumerate(reader, start=2):  # Line 2 (after header)
                        try:
                            # Parse date (supports DD-MM-YYYY or YYYY-MM-DD)
                            date_str = row['Date'].strip()
                            if not date_str:
                                raise ValueError("Date is empty")

                            if '-' in date_str:
                                parts = date_str.split('-')
                                if len(parts[0]) == 4:  # YYYY-MM-DD
                                    trade_date = date_str
                                else:  # DD-MM-YYYY
                                    day, month, year = parts
                                    trade_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                            else:
                                raise ValueError(f"Invalid date format: {date_str} (use DD-MM-YYYY or YYYY-MM-DD)")

                            equity = row['Stock'].strip().upper()
                            if not equity:
                                raise ValueError("Stock symbol is empty")

                            trade_type = row['Type'].strip().upper()
                            if trade_type not in ['BUY', 'SELL']:
                                raise ValueError(f"Type must be BUY or SELL, got: {trade_type}")

                            quantity = int(row['Qty'])
                            if quantity <= 0:
                                raise ValueError(f"Quantity must be positive, got: {quantity}")

                            price_rupees = float(row['Price'])
                            if price_rupees <= 0:
                                raise ValueError(f"Price must be positive, got: {price_rupees}")
                            price_paise = int(price_rupees * 100)

                            brokerage_paise = 0
                            if 'Brokerage' in row and row['Brokerage'].strip():
                                brokerage_rupees = float(row['Brokerage'])
                                if brokerage_rupees < 0:
                                    raise ValueError("Brokerage must be >= 0")
                                brokerage_paise = int(brokerage_rupees * 100)

                            brokerage_override = None
                            if 'BrokerageOverride' in row and row['BrokerageOverride'].strip():
                                override_rupees = float(row['BrokerageOverride'])
                                if override_rupees < 0:
                                    raise ValueError("Brokerage override must be >= 0")
                                brokerage_override = int(override_rupees * 100)

                            mtf_amount_paise = 0
                            if 'MtfAmount' in row and row['MtfAmount'].strip():
                                mtf_amount_rupees = float(row['MtfAmount'])
                                if mtf_amount_rupees < 0:
                                    raise ValueError("MTF amount must be >= 0")
                                mtf_amount_paise = int(mtf_amount_rupees * 100)
                                
                            mtf_rate_ppm = None
                            if 'MtfRate' in row and row['MtfRate'].strip():
                                mtf_rate_percent = float(row['MtfRate'])
                                if mtf_rate_percent < 0:
                                    raise ValueError("MTF rate must be >= 0")
                                mtf_rate_ppm = int(mtf_rate_percent * 10000)

                            notes = row.get('Notes', '').strip()

                            trade_ts = ""
                            if 'TradeTS' in row and row['TradeTS'].strip():
                                trade_ts = row['TradeTS'].strip()
                            if not trade_ts:
                                trade_ts = make_trade_ts(trade_date, "09:15:00")

                            type1_raw = (row.get('Type1') or '').strip()
                            type2_raw = (row.get('Type2') or '').strip()
                            strike_raw = (row.get('Strike') or '').strip()
                            expiry_raw = (row.get('Expiry') or '').strip()

                            if not type1_raw:
                                type1_raw = 'delivery'

                            type1, type2, strike, expiry = normalize_trade_classification(
                                type1_raw,
                                type2_raw,
                                strike_raw,
                                expiry_raw,
                                require_type1=True
                            )

                            type1_norm = (type1 or '').lower()

                            if type1_norm == 'mtf' and trade_type == 'BUY' and mtf_amount_paise <= 0:
                                raise ValueError("MTF amount is required for MTF BUY trades")

                            if type1_norm == 'mtf' and trade_type == 'BUY':
                                trade_amount = quantity * price_paise
                                if mtf_amount_paise > trade_amount:
                                    raise ValueError("MTF amount cannot exceed buy trade amount")

                            # Determine brokerage auto/override
                            if brokerage_override is not None:
                                brokerage_auto = 0
                                brokerage_paise = brokerage_override
                            elif brokerage_paise > 0:
                                # Legacy Brokerage column behaves like override for imports
                                brokerage_auto = 0
                                brokerage_override = brokerage_paise
                            else:
                                try:
                                    brokerage_paise, _rate_ppm = calculate_brokerage_auto(
                                        quantity, price_paise, type1_norm, trade_type
                                    )
                                    brokerage_auto = brokerage_paise
                                except Exception:
                                    raise ValueError("Brokerage override required when rate is not configured")

                            # Insert into database
                            cursor.execute("""
                                INSERT INTO trade_events (
                                    trade_date, equity, trade_type, quantity, price, brokerage,
                                    brokerage_auto, brokerage_override, mtf_amount, mtf_rate_ppm, trade_ts, notes,
                                    type1, type2, strike, expiry, is_active, profile_id
                                )
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
                            """, (
                                trade_date, equity, trade_type, quantity, price_paise, brokerage_paise,
                                brokerage_auto, brokerage_override, mtf_amount_paise, mtf_rate_ppm, trade_ts, notes,
                                type1, type2, strike, expiry, profile_id
                            ))

                            imported_count += 1

                        except Exception as e:
                            skipped_count += 1
                            error_lines.append(f"Line {line_num}: {str(e)}")
                            logger.warning(f"Skipped line {line_num}: {str(e)}")
            
            # Show result
            result_msg = f"\u2705 Imported {imported_count} trades successfully"
            if skipped_count > 0:
                result_msg += f"\\n\\n\u26a0\ufe0f Skipped {skipped_count} rows with errors:\\n"
                result_msg += "\\n".join(error_lines[:10])  # Show first 10 errors
                if len(error_lines) > 10:
                    result_msg += f"\\n... and {len(error_lines) - 10} more errors"
            
            messagebox.showinfo("Import Complete", result_msg)
            self.refresh_records()
            self.update_status(f"\u2705 Imported {imported_count} trades from CSV")
            logger.info(f"CSV import complete: {imported_count} imported, {skipped_count} skipped")
            
        except Exception as e:
            logger.error(f"Failed to import CSV: {str(e)}", exc_info=True)
            messagebox.showerror("Import Failed", f"Failed to import CSV:\\n\\n{str(e)}")
            self.update_status("\u274c CSV import failed")
    
    def export_to_csv(self) -> None:
        """Export current filtered records to CSV."""
        logger.info("Exporting records to CSV")
        
        try:
            # Get timestamp for filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"trades_export_{timestamp}.csv"
            filepath = config.EXPORTS_DIR / filename
            filepath.parent.mkdir(parents=True, exist_ok=True)
            
            rows = self._fetch_filtered_trades_for_export()
            
            if not rows:
                messagebox.showwarning("No Data", "No records to export")
                return
            
            # Write CSV
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([
                    'ID', 'Date', 'TradeTS', 'Stock', 'Type', 'Type1', 'Type2', 'Strike', 'Expiry',
                    'Qty', 'PriceRupees', 'PricePaise',
                    'BrokerageRupees', 'BrokeragePaise', 'BrokerageAutoPaise', 'BrokerageOverridePaise',
                    'MtfAmountRupees', 'MtfAmountPaise', 'MtfRate', 'Notes', 'Status'
                ])

                for row in rows:
                    (trade_id, trade_date, trade_ts, equity, trade_type, type1, type2, strike, expiry,
                     quantity, price_paise, brokerage_paise, brokerage_auto, brokerage_override,
                     mtf_amount_paise, mtf_rate_ppm, notes, is_active) = row

                    display_date = ""
                    if trade_date:
                        year, month, day = trade_date.split('-')
                        display_date = f"{day}-{month}-{year}"

                    expiry_display = ""
                    if expiry:
                        year_e, month_e, day_e = expiry.split('-')
                        expiry_display = f"{day_e}-{month_e}-{year_e}"

                    writer.writerow([
                        trade_id,
                        display_date,
                        trade_ts or "",
                        equity,
                        trade_type,
                        (type1 or "delivery").upper(),
                        type2 or "",
                        "" if strike is None else f"{strike:.2f}",
                        expiry_display,
                        quantity,
                        f"{(price_paise or 0) / 100:.2f}",
                        int(price_paise or 0),
                        f"{(brokerage_paise or 0) / 100:.2f}",
                        int(brokerage_paise or 0),
                        int(brokerage_auto or 0),
                        "" if brokerage_override is None else int(brokerage_override),
                        f"{(mtf_amount_paise or 0) / 100:.2f}",
                        int(mtf_amount_paise or 0),
                        f"{(mtf_rate_ppm / 10000):.2f}" if mtf_rate_ppm is not None else "",
                        notes or "",
                        "Active" if is_active == 1 else "Deleted"
                    ])
            
            logger.info(f"✅ Exported {len(rows)} records to {filepath}")
            self.update_status(f"✅ Exported to {filename}")
            messagebox.showinfo("Export Successful", f"Exported {len(rows)} records to:\n{filepath}")
            
        except Exception as e:
            logger.error(f"CSV export failed: {str(e)}", exc_info=True)
            messagebox.showerror("Export Failed", f"Failed to export CSV:\n{str(e)}")
    
    def export_to_excel(self) -> None:
        """Export current filtered records to Excel."""
        logger.info("Exporting records to Excel")
        
        try:
            # Try importing openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                logger.warning("openpyxl not installed")
                messagebox.showerror(
                    "Missing Dependency",
                    "openpyxl is required for Excel export.\n\n"
                    "Install it with:\npip install openpyxl"
                )
                return
            
            # Get timestamp for filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"trades_export_{timestamp}.xlsx"
            filepath = config.EXPORTS_DIR / filename
            filepath.parent.mkdir(parents=True, exist_ok=True)
            
            rows = self._fetch_filtered_trades_for_export()
            
            if not rows:
                messagebox.showwarning("No Data", "No records to export")
                return
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Trades"
            
            # Header row
            headers = [
                'ID', 'Date', 'TradeTS', 'Stock', 'Type', 'Type1', 'Type2', 'Strike', 'Expiry',
                'Qty', 'PriceRupees', 'PricePaise',
                'BrokerageRupees', 'BrokeragePaise', 'BrokerageAutoPaise', 'BrokerageOverridePaise',
                'MtfAmountRupees', 'MtfAmountPaise', 'MtfRate', 'Notes', 'Status'
            ]
            ws.append(headers)
            
            # Style header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            if ws:  # Type guard for openpyxl worksheet
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for row in rows:
                (trade_id, trade_date, trade_ts, equity, trade_type, type1, type2, strike, expiry,
                 quantity, price_paise, brokerage_paise, brokerage_auto, brokerage_override,
                 mtf_amount_paise, mtf_rate_ppm, notes, is_active) = row

                display_date = ""
                if trade_date:
                    year, month, day = trade_date.split('-')
                    display_date = f"{day}-{month}-{year}"

                expiry_display = ""
                if expiry:
                    year_e, month_e, day_e = expiry.split('-')
                    expiry_display = f"{day_e}-{month_e}-{year_e}"

                ws.append([
                    trade_id,
                    display_date,
                    trade_ts or "",
                    equity,
                    trade_type,
                    (type1 or "delivery").upper(),
                    type2 or "",
                    "" if strike is None else float(strike),
                    expiry_display,
                    quantity,
                    float((price_paise or 0) / 100),
                    int(price_paise or 0),
                    float((brokerage_paise or 0) / 100),
                    int(brokerage_paise or 0),
                    int(brokerage_auto or 0),
                    "" if brokerage_override is None else int(brokerage_override),
                    float((mtf_amount_paise or 0) / 100),
                    int(mtf_amount_paise or 0),
                    float(mtf_rate_ppm / 10000) if mtf_rate_ppm is not None else "",
                    notes or "",
                    "Active" if is_active == 1 else "Deleted"
                ])
            
            # Auto-width columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save
            wb.save(filepath)
            
            logger.info(f"✅ Exported {len(rows)} records to {filepath}")
            self.update_status(f"✅ Exported to {filename}")
            messagebox.showinfo("Export Successful", f"Exported {len(rows)} records to:\n{filepath}")
            
        except Exception as e:
            logger.error(f"Excel export failed: {str(e)}", exc_info=True)
            messagebox.showerror("Export Failed", f"Failed to export Excel:\n{str(e)}")
    
    def backup_database(self) -> None:
        """Create a timestamped backup of the database."""
        logger.info("Creating database backup")
        
        try:
            import shutil
            from tkinter import filedialog
            
            # Generate default filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"trades_backup_{timestamp}.db"
            
            # Ask user where to save
            filepath = filedialog.asksaveasfilename(
                title="Save Database Backup",
                defaultextension=".db",
                initialfile=default_filename,
                filetypes=[("Database files", "*.db"), ("All files", "*.*")]
            )
            
            if not filepath:
                logger.info("Backup cancelled by user")
                return
            
            # Copy database file
            shutil.copy2(str(config.DB_PATH), filepath)
            
            logger.info(f"✅ Database backed up to: {filepath}")
            self.update_status(f"✅ Backup saved: {Path(filepath).name}")
            messagebox.showinfo("Backup Successful", f"Database backed up to:\n{filepath}")
            
        except Exception as e:
            logger.error(f"Backup failed: {str(e)}", exc_info=True)
            messagebox.showerror("Backup Failed", f"Failed to backup database:\n{str(e)}")
    
    def restore_database(self) -> None:
        """Restore database from a backup file."""
        logger.info("Restoring database from backup")
        
        try:
            import shutil
            from tkinter import filedialog
            
            # Warn user
            result = messagebox.askyesno(
                "Confirm Restore",
                "⚠️ WARNING: This will replace your current database!\n\n"
                "All current trades will be replaced with the backup data.\n\n"
                "It's recommended to create a backup of your current database first.\n\n"
                "Do you want to continue?"
            )
            
            if not result:
                logger.info("Restore cancelled by user")
                return
            
            # Ask user to select backup file
            filepath = filedialog.askopenfilename(
                title="Select Backup File to Restore",
                filetypes=[("Database files", "*.db"), ("All files", "*.*")]
            )
            
            if not filepath:
                logger.info("Restore cancelled - no file selected")
                return
            
            # Create a safety backup of current database
            safety_backup = config.DB_BACKUP_DIR / f"trades_before_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            shutil.copy2(str(config.DB_PATH), str(safety_backup))
            logger.info(f"Created safety backup: {safety_backup}")
            
            # Restore from backup
            shutil.copy2(filepath, str(config.DB_PATH))
            
            logger.info(f"✅ Database restored from: {filepath}")
            self.update_status("✅ Database restored")
            messagebox.showinfo(
                "Restore Successful",
                f"Database restored from:\n{filepath}\n\n"
                f"Safety backup created:\n{safety_backup}\n\n"
                "Refreshing records..."
            )
            
            # Refresh display
            self.refresh_records()
            
        except Exception as e:
            logger.error(f"Restore failed: {str(e)}", exc_info=True)
            messagebox.showerror("Restore Failed", f"Failed to restore database:\n{str(e)}")


class EditTradeDialog:
    """Dialog for editing trade details."""
    
    def __init__(self, parent, trade_id: int, trade_data: tuple, refresh_callback: Callable, status_callback: Callable):
        self.trade_id = trade_id
        self.refresh_callback = refresh_callback
        self.update_status = status_callback
        
        # Create dialog window
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(f"Edit Trade #{trade_id}")
        self.dialog.geometry("600x650")
        self.dialog.resizable(False, False)  # Prevent accidental resizing
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Unpack trade data
        (trade_date, trade_ts, equity, trade_type, type1, type2, strike, expiry,
         quantity, price_paise, brokerage_paise, _brokerage_auto, brokerage_override,
         mtf_amount_paise, mtf_rate_ppm, notes, _) = trade_data
        
        # Convert for display
        price_rupees = price_paise / 100
        brokerage_rupees = brokerage_paise / 100
        mtf_amount_rupees = (mtf_amount_paise or 0) / 100
        self.trade_ts = trade_ts
        year, month, day = trade_date.split('-')
        display_date = f"{day}-{month}-{year}"
        
        # Create form
        main_frame = ttk.Frame(self.dialog, padding="20")
        main_frame.pack(fill='both', expand=True)
        
        ttk.Label(main_frame, text=f"EDIT TRADE #{trade_id}", font=('Consolas', 12, 'bold')).grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        row = 1
        
        # Trade ID (readonly display)
        ttk.Label(main_frame, text="Trade ID:", font=('Consolas', 10)).grid(row=row, column=0, sticky='e', padx=5, pady=8)
        id_display = ttk.Label(main_frame, text=f"#{trade_id}", font=('Consolas', 10, 'bold'), foreground='blue')
        id_display.grid(row=row, column=1, sticky='w', padx=5, pady=8)
        row += 1
        
        # Date
        ttk.Label(main_frame, text="Date:", font=('Arial', 10)).grid(row=row, column=0, sticky='e', padx=5, pady=8)
        self.date_entry = ttk.Entry(main_frame, width=25)
        self.date_entry.grid(row=row, column=1, sticky='w', padx=5, pady=8)
        self.date_entry.insert(0, display_date)
        row += 1
        
        # Equity
        ttk.Label(main_frame, text="Stock:", font=('Arial', 10)).grid(row=row, column=0, sticky='e', padx=5, pady=8)
        self.equity_entry = ttk.Entry(main_frame, width=25)
        self.equity_entry.grid(row=row, column=1, sticky='w', padx=5, pady=8)
        self.equity_entry.insert(0, equity)
        row += 1
        
        # Trade Type
        ttk.Label(main_frame, text="Type:", font=('Arial', 10)).grid(row=row, column=0, sticky='e', padx=5, pady=8)
        self.trade_type_var = tk.StringVar(value=trade_type)
        type_frame = ttk.Frame(main_frame)
        type_frame.grid(row=row, column=1, sticky='w', padx=5, pady=8)
        ttk.Radiobutton(type_frame, text="BUY", variable=self.trade_type_var, value='BUY').pack(side='left', padx=5)
        ttk.Radiobutton(type_frame, text="SELL", variable=self.trade_type_var, value='SELL').pack(side='left', padx=5)
        row += 1

        # Type1 (classification)
        ttk.Label(main_frame, text="Type1:", font=('Arial', 10)).grid(row=row, column=0, sticky='e', padx=5, pady=8)
        self.type1_var = tk.StringVar(value=(type1 or "DELIVERY").upper())
        self.type1_entry = ttk.Combobox(
            main_frame,
            textvariable=self.type1_var,
            values=['INTRADAY', 'DELIVERY', 'MTF', 'FUTURES', 'OPTIONS'],
            state='readonly',
            width=23
        )
        self.type1_entry.grid(row=row, column=1, sticky='w', padx=5, pady=8)
        self.type1_entry.bind('<<ComboboxSelected>>', lambda _e: self.update_derivative_fields())
        row += 1

        # Type2 (options only)
        self.type2_label = ttk.Label(main_frame, text="Type2:", font=('Arial', 10))
        self.type2_label.grid(row=row, column=0, sticky='e', padx=5, pady=8)
        self.type2_var = tk.StringVar(value=type2 or "")
        self.type2_entry = ttk.Combobox(
            main_frame,
            textvariable=self.type2_var,
            values=['CE', 'PE'],
            state='readonly',
            width=23
        )
        self.type2_entry.grid(row=row, column=1, sticky='w', padx=5, pady=8)
        row += 1

        # Strike (options only)
        self.strike_label = ttk.Label(main_frame, text="Strike:", font=('Arial', 10))
        self.strike_label.grid(row=row, column=0, sticky='e', padx=5, pady=8)
        self.strike_entry = ttk.Entry(main_frame, width=25)
        self.strike_entry.grid(row=row, column=1, sticky='w', padx=5, pady=8)
        if strike is not None:
            self.strike_entry.insert(0, f"{strike:.2f}")
        row += 1

        # Expiry (options/futures)
        self.expiry_label = ttk.Label(main_frame, text="Expiry:", font=('Arial', 10))
        self.expiry_label.grid(row=row, column=0, sticky='e', padx=5, pady=8)
        self.expiry_entry = ttk.Entry(main_frame, width=25)
        self.expiry_entry.grid(row=row, column=1, sticky='w', padx=5, pady=8)
        if expiry:
            year_e, month_e, day_e = expiry.split('-')
            self.expiry_entry.insert(0, f"{day_e}-{month_e}-{year_e}")
        row += 1
        
        # Quantity
        ttk.Label(main_frame, text="Quantity:", font=('Arial', 10)).grid(row=row, column=0, sticky='e', padx=5, pady=8)
        self.quantity_entry = ttk.Entry(main_frame, width=25)
        self.quantity_entry.grid(row=row, column=1, sticky='w', padx=5, pady=8)
        self.quantity_entry.insert(0, str(quantity))
        row += 1
        
        # Price
        ttk.Label(main_frame, text="Price (₹):", font=('Arial', 10)).grid(row=row, column=0, sticky='e', padx=5, pady=8)
        self.price_entry = ttk.Entry(main_frame, width=25)
        self.price_entry.grid(row=row, column=1, sticky='w', padx=5, pady=8)
        self.price_entry.insert(0, f"{price_rupees:.2f}")
        row += 1
        
        # Brokerage
        ttk.Label(main_frame, text="Brokerage (₹):", font=('Arial', 10)).grid(row=row, column=0, sticky='e', padx=5, pady=8)
        self.brokerage_entry = ttk.Entry(main_frame, width=25)
        self.brokerage_entry.grid(row=row, column=1, sticky='w', padx=5, pady=8)
        self.brokerage_entry.insert(0, f"{brokerage_rupees:.2f}")
        row += 1

        # Brokerage override flag (display)
        self.override_brokerage_var = tk.BooleanVar(value=brokerage_override is not None)
        self.override_brokerage_check = ttk.Checkbutton(
            main_frame,
            text="Override brokerage (edit value)",
            variable=self.override_brokerage_var
        )
        self.override_brokerage_check.grid(row=row, column=1, sticky='w', padx=5, pady=(0, 8))
        row += 1

        # MTF amount
        ttk.Label(main_frame, text="MTF Amount (₹):", font=('Arial', 10)).grid(row=row, column=0, sticky='e', padx=5, pady=8)
        self.mtf_amount_entry = ttk.Entry(main_frame, width=25)
        self.mtf_amount_entry.grid(row=row, column=1, sticky='w', padx=5, pady=8)
        self.mtf_amount_entry.insert(0, f"{mtf_amount_rupees:.2f}")
        row += 1
        
        # MTF rate
        ttk.Label(main_frame, text="MTF Rate (%):", font=('Arial', 10)).grid(row=row, column=0, sticky='e', padx=5, pady=8)
        self.mtf_rate_entry = ttk.Entry(main_frame, width=25)
        self.mtf_rate_entry.grid(row=row, column=1, sticky='w', padx=5, pady=8)
        if mtf_rate_ppm is not None:
            self.mtf_rate_entry.insert(0, f"{(mtf_rate_ppm / 10000):.2f}")
        row += 1
        
        # Notes
        ttk.Label(main_frame, text="Notes:", font=('Arial', 10)).grid(row=row, column=0, sticky='ne', padx=5, pady=8)
        self.notes_entry = tk.Text(main_frame, width=30, height=3, font=('Arial', 10))
        self.notes_entry.grid(row=row, column=1, sticky='w', padx=5, pady=8)
        self.notes_entry.insert('1.0', notes)
        row += 1

        self.update_derivative_fields()
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=row, column=0, columnspan=2, pady=20)
        
        ttk.Button(button_frame, text="Cancel", command=self.dialog.destroy, width=15).pack(side='left', padx=10)
        ttk.Button(button_frame, text="Save Changes", command=self.save_changes, width=15).pack(side='left', padx=10)

    def update_derivative_fields(self) -> None:
        """Enable/disable derivative fields based on Type1 selection."""
        type1 = self.type1_var.get().strip().lower()
        is_options = type1 == 'options'
        is_futures = type1 == 'futures'

        def show_option_fields() -> None:
            self.type2_label.grid()
            self.type2_entry.grid()
            self.strike_label.grid()
            self.strike_entry.grid()
            self.expiry_label.grid()
            self.expiry_entry.grid()

        def show_futures_fields() -> None:
            self.type2_label.grid_remove()
            self.type2_entry.grid_remove()
            self.strike_label.grid_remove()
            self.strike_entry.grid_remove()
            self.expiry_label.grid()
            self.expiry_entry.grid()

        def hide_all_derivative_fields() -> None:
            self.type2_label.grid_remove()
            self.type2_entry.grid_remove()
            self.strike_label.grid_remove()
            self.strike_entry.grid_remove()
            self.expiry_label.grid_remove()
            self.expiry_entry.grid_remove()

        if is_options:
            show_option_fields()
        elif is_futures:
            self.type2_var.set('')
            self.strike_entry.delete(0, tk.END)
            show_futures_fields()
        else:
            self.type2_var.set('')
            self.strike_entry.delete(0, tk.END)
            self.expiry_entry.delete(0, tk.END)
            hide_all_derivative_fields()
    
    def save_changes(self) -> None:
        """Save edited trade to database."""
        try:
            # Get values
            date_str = self.date_entry.get().strip()
            day, month, year = date_str.split('-')
            trade_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
            time_part = "09:15:00"
            if getattr(self, 'trade_ts', None) and ' ' in self.trade_ts:
                time_part = self.trade_ts.split(' ', 1)[1]
            trade_ts = make_trade_ts(trade_date, time_part)
            
            equity = self.equity_entry.get().strip().upper()
            trade_type = self.trade_type_var.get()
            quantity = int(self.quantity_entry.get())

            type1, type2, strike, expiry = normalize_trade_classification(
                self.type1_var.get(),
                self.type2_var.get(),
                self.strike_entry.get(),
                self.expiry_entry.get(),
                require_type1=True
            )
            
            price_rupees = float(self.price_entry.get())
            price_paise = int(price_rupees * 100)
            
            brokerage_rupees = float(self.brokerage_entry.get())
            brokerage_paise = int(brokerage_rupees * 100)
            if self.override_brokerage_var.get():
                brokerage_override = brokerage_paise
                brokerage_auto = 0
            else:
                try:
                    type1_for_calc = type1 or "delivery"
                    brokerage_paise, _rate_ppm = calculate_brokerage_auto(
                        quantity, price_paise, type1_for_calc, trade_type
                    )
                    brokerage_auto = brokerage_paise
                    brokerage_override = None
                except Exception:
                    messagebox.showerror(
                        "Brokerage Required",
                        "Brokerage rate not configured. Enable override and enter brokerage manually."
                    )
                    return

            mtf_amount_rupees = float(self.mtf_amount_entry.get() or 0)
            mtf_amount_paise = int(mtf_amount_rupees * 100)

            mtf_rate_val = self.mtf_rate_entry.get().strip()
            mtf_rate_ppm = None
            if mtf_rate_val:
                mtf_rate_ppm = int(float(mtf_rate_val) * 10000)

            if type1 == 'mtf' and trade_type == 'BUY':
                if mtf_amount_paise <= 0:
                    messagebox.showerror("Invalid MTF Amount", "MTF amount is required for MTF BUY trades")
                    return
                trade_amount = quantity * price_paise
                if mtf_amount_paise > trade_amount:
                    messagebox.showerror("Invalid MTF Amount", "MTF amount cannot exceed buy trade amount")
                    return
            
            notes = self.notes_entry.get('1.0', 'end-1c').strip()

            logger.info(f"Updating trade ID {self.trade_id}: {trade_type} {quantity} {equity} @ ₹{price_rupees:.2f}")
            
            # Update database
            conn = sqlite3.connect(str(config.DB_PATH))
            c = conn.cursor()
            c.execute("""
                UPDATE trade_events
                SET trade_date = ?, trade_ts = ?, equity = ?, trade_type = ?, type1 = ?, type2 = ?,
                    strike = ?, expiry = ?, quantity = ?, price = ?, brokerage = ?,
                    brokerage_auto = ?, brokerage_override = ?, mtf_amount = ?, mtf_rate_ppm = ?, notes = ?
                WHERE id = ?
            """, (
                trade_date, trade_ts, equity, trade_type, type1, type2, strike, expiry,
                quantity, price_paise, brokerage_paise, brokerage_auto, brokerage_override,
                mtf_amount_paise, mtf_rate_ppm, notes, self.trade_id
            ))
            conn.commit()
            conn.close()
            
            logger.info(f"✅ Trade ID {self.trade_id} updated successfully")
            self.update_status(f"✅ Trade #{self.trade_id} updated")
            messagebox.showinfo("Success", "Trade updated successfully")
            
            # Close dialog and refresh
            self.dialog.destroy()
            self.refresh_callback()
            
        except Exception as e:
            logger.error(f"Failed to update trade: {str(e)}", exc_info=True)
            messagebox.showerror("Error", f"Failed to update trade:\n{str(e)}")
