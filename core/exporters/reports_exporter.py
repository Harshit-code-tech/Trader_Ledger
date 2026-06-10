import csv
from datetime import datetime
from pathlib import Path
import config
from core.logger import get_logger

logger = get_logger('core.exporters')

def export_audit_csv(matches: list, trades_by_id: dict, trade_ts_map: dict, remainder_flags: dict) -> str:
    """Export match-level audit details to CSV. Returns the filepath."""
    if not matches:
        raise ValueError("No match-level data to export.")

    config.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = config.EXPORTS_DIR / f"audit_{timestamp}.csv"

    def rupees(paise: int, absolute: bool = False) -> str:
        value = abs(paise) if absolute else paise
        return f"{value / 100:.2f}"

    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            "SellID", "BuyID", "Equity", "Type1", "Type2", "Strike", "Expiry",
            "MatchedQty", "BuyDate", "BuyTS", "SellDate", "SellTS",
            "BuyBrokerageAutoPaise", "BuyBrokerageOverridePaise", "BuyBrokerageEffectivePaise",
            "SellBrokerageAutoPaise", "SellBrokerageOverridePaise", "SellBrokerageEffectivePaise",
            "BuyPrice", "SellPrice", "BuyCost", "SellValue",
            "BuyBrokerageAlloc", "SellBrokerageAlloc",
            "MatchedBuyTotal", "MatchedSellTotal", "GrossPnL",
            "MatchedMtfAmount", "HoldingDays", "MtfInterest", "NetPnL",
            "AllocationRemainderApplied",
            "BuyPricePaise", "SellPricePaise", "BuyCostPaise", "SellValuePaise",
            "BuyBrokerageAllocPaise", "SellBrokerageAllocPaise",
            "MatchedBuyTotalPaise", "MatchedSellTotalPaise", "GrossPnLPaise",
            "MatchedMtfAmountPaise", "MtfInterestPaise", "NetPnLPaise"
        ])

        for idx, match in enumerate(matches):
            buy = trades_by_id.get(match['buy_id'])
            sell = trades_by_id.get(match['sell_id'])
            if not buy or not sell:
                continue

            buy_date = buy.get('trade_date') or ""
            sell_date = sell.get('trade_date') or ""
            buy_ts = trade_ts_map.get(match['buy_id']) or (f"{buy_date} 09:15:00" if buy_date else "")
            sell_ts = trade_ts_map.get(match['sell_id']) or (f"{sell_date} 09:15:00" if sell_date else "")

            gross_pnl = match.get('gross_pnl', match.get('realized_pnl', 0))
            net_pnl = match.get('net_pnl', gross_pnl)
            buy_brokerage_override = buy.get('brokerage_override')
            sell_brokerage_override = sell.get('brokerage_override')

            writer.writerow([
                match['sell_id'], match['buy_id'], sell.get('equity', ''),
                sell.get('type1', ''), sell.get('type2', '') or '',
                "" if sell.get('strike') is None else str(sell.get('strike')),
                sell.get('expiry', '') or '', match['matched_quantity'],
                buy_date, buy_ts, sell_date, sell_ts,
                int(buy.get('brokerage_auto', 0) or 0),
                "" if buy_brokerage_override is None else int(buy_brokerage_override),
                int(buy.get('brokerage', 0) or 0),
                int(sell.get('brokerage_auto', 0) or 0),
                "" if sell_brokerage_override is None else int(sell_brokerage_override),
                int(sell.get('brokerage', 0) or 0),
                rupees(int(buy.get('price', 0)), absolute=True),
                rupees(int(sell.get('price', 0)), absolute=True),
                rupees(match['buy_cost'], absolute=True), rupees(match['sell_value'], absolute=True),
                rupees(match['buy_brokerage_alloc'], absolute=True), rupees(match['sell_brokerage_alloc'], absolute=True),
                rupees(match.get('matched_buy_total', match['buy_cost'] + match['buy_brokerage_alloc']), absolute=True),
                rupees(match.get('matched_sell_total', match['sell_value'] - match['sell_brokerage_alloc']), absolute=True),
                rupees(gross_pnl), rupees(int(match.get('matched_mtf_amount', 0)), absolute=True),
                match.get('holding_days', 0), rupees(int(match.get('mtf_interest', 0)), absolute=True),
                rupees(net_pnl), remainder_flags.get(idx, "NONE"),
                int(buy.get('price', 0) or 0), int(sell.get('price', 0) or 0),
                int(match['buy_cost']), int(match['sell_value']),
                int(match['buy_brokerage_alloc']), int(match['sell_brokerage_alloc']),
                int(match.get('matched_buy_total', match['buy_cost'] + match['buy_brokerage_alloc'])),
                int(match.get('matched_sell_total', match['sell_value'] - match['sell_brokerage_alloc'])),
                int(gross_pnl), int(match.get('matched_mtf_amount', 0) or 0),
                int(match.get('mtf_interest', 0) or 0), int(net_pnl)
            ])
            
    logger.info(f"Audit CSV exported successfully to {filepath}")
    return str(filepath)


def export_report_csv(data: dict) -> str:
    """Export current filtered report to CSV. Returns the filepath."""
    config.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = config.EXPORTS_DIR / f"report_{timestamp}.csv"

    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        writer.writerow(["Filters"])
        writer.writerow(["Equities", data['filters']['equities']])
        writer.writerow(["From Date", data['filters']['from_date'] or "-"])
        writer.writerow(["To Date", data['filters']['to_date'] or "-"])
        writer.writerow(["Type1", data['filters']['type1'] or "All"])
        writer.writerow(["Expiry Month", data['filters']['expiry_month'] or "-"])
        writer.writerow(["Include Open Positions", str(data['filters']['include_open_positions'])])
        writer.writerow([])

        writer.writerow(["Summary"])
        writer.writerow(["Total Profit", data['summary']['total_profit']])
        writer.writerow(["Total Loss", data['summary']['total_loss']])
        writer.writerow(["Net P/L", data['summary']['net_pnl']])
        writer.writerow(["Win/Loss Ratio", data['summary']['win_loss_ratio']])
        writer.writerow(["Win Rate", data['summary']['win_rate']])
        writer.writerow(["Profit Factor", data['summary']['profit_factor']])
        writer.writerow(["Avg Win", data['summary']['avg_win']])
        writer.writerow(["Avg Loss", data['summary']['avg_loss']])
        writer.writerow(["Expectancy", data['summary']['expectancy']])
        writer.writerow(["Avg Holding Period (days)", data['summary']['avg_holding_days']])
        writer.writerow(["Median Holding Period (days)", data['summary']['median_holding_days']])
        writer.writerow(["Max Drawdown", data['summary']['max_drawdown']])
        writer.writerow([])

        writer.writerow([f"{data['period']['type']} P/L"])
        writer.writerow(["Period", "Profit", "Loss", "Net P/L", "Running Total"])
        for row in data['period']['rows']:
            writer.writerow(row)
        writer.writerow([])

        writer.writerow(["Equity-wise Summary"])
        writer.writerow(["Equity", "Closed P/L", "Open P/L", "Total"])
        for row in data['equity_summary']:
            writer.writerow(row)
        writer.writerow([])

        if data['open_positions']:
            writer.writerow(["Open Positions"])
            writer.writerow([
                "Equity", "Type1", "Type2", "Strike", "Expiry",
                "Holding Days", "Status", "Qty", "Avg Price", "Unrealized P/L"
            ])
            for row in data['open_positions']:
                writer.writerow(row)

    logger.info(f"Report CSV exported successfully to {filepath}")
    return str(filepath)


def export_report_excel(data: dict) -> str:
    """Export current filtered report to Excel (xlsx). Returns the filepath."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        raise ImportError("openpyxl is required for Excel export.\n\nInstall it with:\npip install openpyxl")

    config.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = config.EXPORTS_DIR / f"report_{timestamp}.xlsx"

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    bold = Font(bold=True)

    ws_summary.append(["Filters"])
    ws_summary["A1"].font = bold
    ws_summary.append(["Equities", data['filters']['equities']])
    ws_summary.append(["From Date", data['filters']['from_date'] or "-"])
    ws_summary.append(["To Date", data['filters']['to_date'] or "-"])
    ws_summary.append(["Type1", data['filters']['type1'] or "All"])
    ws_summary.append(["Expiry Month", data['filters']['expiry_month'] or "-"])
    ws_summary.append(["Include Open Positions", str(data['filters']['include_open_positions'])])
    ws_summary.append([])
    ws_summary.append(["Summary"])
    ws_summary["A7"].font = bold
    ws_summary.append(["Total Profit", data['summary']['total_profit']])
    ws_summary.append(["Total Loss", data['summary']['total_loss']])
    ws_summary.append(["Net P/L", data['summary']['net_pnl']])
    ws_summary.append(["Win/Loss Ratio", data['summary']['win_loss_ratio']])
    ws_summary.append(["Win Rate", data['summary']['win_rate']])
    ws_summary.append(["Profit Factor", data['summary']['profit_factor']])
    ws_summary.append(["Avg Win", data['summary']['avg_win']])
    ws_summary.append(["Avg Loss", data['summary']['avg_loss']])
    ws_summary.append(["Expectancy", data['summary']['expectancy']])
    ws_summary.append(["Avg Holding Period (days)", data['summary']['avg_holding_days']])
    ws_summary.append(["Median Holding Period (days)", data['summary']['median_holding_days']])
    ws_summary.append(["Max Drawdown", data['summary']['max_drawdown']])

    ws_period = wb.create_sheet(title=f"{data['period']['type']} PnL")
    ws_period.append(["Period", "Profit", "Loss", "Net P/L", "Running Total"])
    for cell in ws_period[1]:
        cell.font = bold
    for row in data['period']['rows']:
        ws_period.append(list(row))

    ws_equity = wb.create_sheet(title="Equity Summary")
    ws_equity.append(["Equity", "Closed P/L", "Open P/L", "Total"])
    for cell in ws_equity[1]:
        cell.font = bold
    for row in data['equity_summary']:
        ws_equity.append(list(row))

    if data['open_positions']:
        ws_open = wb.create_sheet(title="Open Positions")
        ws_open.append([
            "Equity", "Type1", "Type2", "Strike", "Expiry",
            "Holding Days", "Status", "Qty", "Avg Price", "Unrealized P/L"
        ])
        for cell in ws_open[1]:
            cell.font = bold
        for row in data['open_positions']:
            ws_open.append(list(row))

    wb.save(filepath)
    logger.info(f"Excel report exported successfully to {filepath}")
    return str(filepath)


def print_report_html(data: dict) -> str:
    """Generate a print-friendly HTML report. Returns the filepath."""
    config.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = config.EXPORTS_DIR / f"report_{timestamp}.html"

    period_type = data['period']['type']

    html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>P/L Report - {datetime.now().strftime("%d %b %Y %I:%M %p")}</title>
    <style>
        @media print {{ @page {{ margin: 1cm; }} body {{ margin: 0; }} }}
        body {{ font-family: Consolas, 'Courier New', monospace; max-width: 1000px; margin: 20px auto; padding: 20px; background: white; }}
        h1 {{ text-align: center; color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
        .timestamp {{ text-align: center; color: #7f8c8d; margin-bottom: 30px; }}
        .summary {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 20px; margin-bottom: 30px; }}
        .card {{ border: 2px solid #ecf0f1; border-radius: 8px; padding: 15px; text-align: center; }}
        .card-title {{ font-size: 12px; color: #7f8c8d; margin-bottom: 10px; }}
        .card-value {{ font-size: 24px; font-weight: bold; }}
        .profit {{ color: #27ae60; }}
        .loss {{ color: #e74c3c; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
        th {{ background: #34495e; color: white; padding: 12px; text-align: left; font-weight: bold; }}
        td {{ padding: 10px; border-bottom: 1px solid #ecf0f1; }}
        tr:hover {{ background: #f8f9fa; }}
        .footer {{ margin-top: 40px; text-align: center; color: #7f8c8d; font-size: 12px; border-top: 1px solid #ecf0f1; padding-top: 20px; }}
    </style>
</head>
<body>
    <h1>📊 PROFIT/LOSS REPORT</h1>
    <div class="timestamp">Generated on {datetime.now().strftime("%d %b %Y at %I:%M %p")}</div>
    <div class="summary">
        <div class="card"><div class="card-title">Total Profit</div><div class="card-value profit">{data['summary']['total_profit']}</div></div>
        <div class="card"><div class="card-title">Total Loss</div><div class="card-value loss">{data['summary']['total_loss']}</div></div>
        <div class="card"><div class="card-title">Net P/L</div><div class="card-value">{data['summary']['net_pnl']}</div></div>
    </div>
    <h2>{period_type} P/L Breakdown</h2>
    <table>
        <thead><tr><th>Period</th><th>Profit</th><th>Loss</th><th>Net P/L</th><th>Running Total</th></tr></thead>
        <tbody>
"""

    for row in data['period']['rows']:
        if len(row) >= 5:
            period_name, profit_str, loss_str, net_str, running_str = row
            row_class = 'profit' if '₹' in net_str and '-' not in net_str else 'loss' if '-' in net_str else ''
            
            html_content += f"""
            <tr>
                <td>{period_name}</td>
                <td class="profit">{profit_str}</td>
                <td class="loss">{loss_str}</td>
                <td class="{row_class}">{net_str}</td>
                <td class="{'profit' if '-' not in running_str else 'loss'}">{running_str}</td>
            </tr>"""

    html_content += """
        </tbody>
    </table>
    <div class="footer">
        <p>Trader Ledger - FIFO-based P/L Calculation System</p>
        <p>This report was automatically generated. Please verify all figures.</p>
    </div>
</body>
</html>
"""

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(html_content)

    logger.info(f"HTML Report exported successfully to {filepath}")
    return str(filepath)
